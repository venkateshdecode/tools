import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\JulianeMatussek\Documents\IcAt\Fruchtzwerge\output\IcAt_Overview_Final.xlsx')
ws_assets = wb['Assets']
ws_reordered = wb['Reordered']

print('=== ASSETS SHEET ===')
print('Row | factor | factorgroup | ID | Language | Yellow?')
for row in range(2, min(15, ws_assets.max_row + 1)):
    factor = ws_assets.cell(row, 1).value
    factorgroup = ws_assets.cell(row, 2).value
    id_val = ws_assets.cell(row, 3).value
    lang = ws_assets.cell(row, 4).value
    fill = ws_assets.cell(row, 4).fill
    is_yellow = 'FFFF00' in str(fill.start_color.rgb) if fill.start_color else False
    print(f'{row} | {factor} | {factorgroup} | {id_val} | {lang} | {is_yellow}')

print('\n=== REORDERED SHEET ===')
print('Row | Group | factor | factorgroup | ID | Language | Yellow?')
for row in range(2, min(15, ws_reordered.max_row + 1)):
    group = ws_reordered.cell(row, 1).value
    factor = ws_reordered.cell(row, 2).value
    factorgroup = ws_reordered.cell(row, 3).value
    id_val = ws_reordered.cell(row, 4).value
    lang = ws_reordered.cell(row, 5).value
    fill = ws_reordered.cell(row, 5).fill
    is_yellow = 'FFFF00' in str(fill.start_color.rgb) if fill.start_color else False
    print(f'{row} | {group} | {factor} | {factorgroup} | {id_val} | {lang} | {is_yellow}')
