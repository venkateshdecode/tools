import os
import re
import shutil
import zipfile
from pathlib import Path
from collections import defaultdict
from typing import Dict, Set, List, Tuple
import streamlit as st
import pandas as pd
import tempfile

# Import helper functions from helper.py to reuse exact same logic
from helper import (
    ALLOWED_IMAGE_EXTENSIONS,
    ALLOWED_VIDEO_EXTENSIONS,
    ALLOWED_AUDIO_EXTENSIONS,
    ALLOWED_TEXT_EXTENSIONS,
    NON_IMAGE_EXTENSIONS,
    get_cleaned_filename_without_brand,
    extract_brand
)


def extract_uploaded_zip(uploaded_file):
    """Extract uploaded zip file to temporary directory"""
    temp_dir = tempfile.mkdtemp()
    with zipfile.ZipFile(uploaded_file, 'r') as zip_ref:
        zip_ref.extractall(temp_dir)
    return temp_dir


def read_zip_structure(temp_dir):
    """
    Analyze the extracted zip structure and identify:
    - processed_files folder
    - reports folder (with Excel and PDF)
    """
    temp_path = Path(temp_dir)

    structure = {
        'processed_files': [],
        'excel_report': None,
        'pdf_report': None,
        'excel_data': None,
        'excel_path': None,
        'processed_files_dir': None
    }

    # Look for processed_files folder
    processed_files_dir = temp_path / "processed_files"
    if processed_files_dir.exists() and processed_files_dir.is_dir():
        all_files = [f for f in processed_files_dir.iterdir() if f.is_file()]
        structure['processed_files'] = all_files
        structure['processed_files_dir'] = processed_files_dir

    # Look for reports folder
    reports_dir = temp_path / "reports"
    if reports_dir.exists() and reports_dir.is_dir():
        # Find Excel report - search by prefix "Brand_Assets_Overview"
        excel_report = None
        for file in reports_dir.iterdir():
            if file.is_file() and file.name.startswith("Brand_Assets_Overview") and file.suffix.lower() in ['.xlsx', '.xls']:
                excel_report = file
                break

        if excel_report:
            structure['excel_report'] = excel_report
            structure['excel_path'] = excel_report
            # Read Excel data
            try:
                df_assets = pd.read_excel(excel_report, sheet_name='Assets')
                df_reordered = pd.read_excel(excel_report, sheet_name='Reordered')
                structure['excel_data'] = {'Assets': df_assets, 'Reordered': df_reordered}
            except Exception as e:
                st.warning(f"Could not read Excel file: {e}")

        # Find PDF report - search by prefix "Brand_Assets_Report"
        pdf_report = None
        for file in reports_dir.iterdir():
            if file.is_file() and file.name.startswith("Brand_Assets_Report") and file.suffix.lower() == '.pdf':
                pdf_report = file
                break

        if pdf_report:
            structure['pdf_report'] = pdf_report

    return structure


def extract_metadata_from_excel(excel_data: Dict) -> Dict:
    """
    Extract critical metadata from the existing Excel file:
    - marken_index (brand number to brand name mapping)
    - file_to_factorgroup (filename to folder name mapping)
    - existing files
    """
    df_assets = excel_data['Assets']

    metadata = {
        'marken_index': {},
        'file_to_factorgroup': {},
        'existing_files': set(),
        'folder_names': set()
    }

    # Extract existing files and their factorgroups
    for _, row in df_assets.iterrows():
        file_id = str(row['ID'])
        factorgroup = str(row['factorgroup'])
        factor = str(row['factor'])

        metadata['existing_files'].add(file_id)
        metadata['file_to_factorgroup'][file_id] = factorgroup
        metadata['folder_names'].add(factorgroup)

        # Extract brand number from ID (first 2 digits)
        match = re.match(r'(\d{2})B\d{2}([a-zA-Z]+)', file_id)
        if match:
            brand_num, brand_name = match.groups()
            metadata['marken_index'][brand_name.lower()] = brand_num

    return metadata


def find_duplicate_pngs(processed_files: List[Path]) -> Set[str]:
    """
    Find PNG files that have the same base name as video/audio files.
    These PNGs should be ignored as they're duplicates.
    """
    video_audio_exts = ALLOWED_VIDEO_EXTENSIONS | ALLOWED_AUDIO_EXTENSIONS
    png_to_ignore = set()

    # Group files by base name (without extension)
    files_by_basename = defaultdict(list)
    for file_path in processed_files:
        basename = file_path.stem
        ext = file_path.suffix.lower()
        files_by_basename[basename].append((file_path.name, ext))

    # Check for PNG files with same name as video/audio files
    for basename, files in files_by_basename.items():
        extensions = [ext for _, ext in files]
        has_video_audio = any(ext in video_audio_exts for ext in extensions)
        has_png = '.png' in extensions

        if has_video_audio and has_png:
            # Found a PNG with same name as video/audio - mark PNG for ignoring
            png_files = [name for name, ext in files if ext == '.png']
            png_to_ignore.update(png_files)

    return png_to_ignore


def analyze_processed_files(processed_files: List[Path], metadata: Dict, png_to_ignore: Set[str]):
    """
    Analyze all processed files and organize them using the EXACT same structure as helper.py
    Returns: renamed_files_by_folder_and_marke, file_to_factorgroup
    """
    from helper import extract_brand, get_cleaned_filename_without_brand

    renamed_files_by_folder_and_marke = defaultdict(lambda: defaultdict(list))
    file_to_factorgroup = {}
    marken_index = metadata['marken_index']

    # Reverse the marken_index to get brand_number -> brand_name
    nummer_zu_marke = {v: k for k, v in marken_index.items()}

    for file_path in processed_files:
        if file_path.name in png_to_ignore:
            continue

        filename = file_path.name
        name = file_path.stem
        file_ext = file_path.suffix.lower()

        # Parse filename to extract metadata: 01B11brandname01description.ext
        match = re.match(r"(\d{2})B(\d{2})([a-z0-9]+)", name, re.IGNORECASE)
        if not match:
            continue

        markennummer, blocknummer, marke_from_name = match.groups()

        # CRITICAL FIX: Use the brand name from marken_index based on brand NUMBER
        # This ensures consistency with helper.py
        marke = nummer_zu_marke.get(markennummer, marke_from_name.lower())

        # Try to get factorgroup from existing metadata first
        if name in metadata['file_to_factorgroup']:
            factorgroup = metadata['file_to_factorgroup'][name]
        else:
            # New file - try to infer folder from other files with same block number
            # Look for existing files with same block number
            inferred_folder = None
            for existing_id, existing_folder in metadata['file_to_factorgroup'].items():
                if f"B{blocknummer}" in existing_id:
                    inferred_folder = existing_folder
                    break

            if inferred_folder:
                factorgroup = inferred_folder
            else:
                # Default fallback
                factorgroup = f"{blocknummer}{marke}"

        # Store mapping
        file_to_factorgroup[filename] = factorgroup

        # Add to renamed_files structure (folder_name -> brand -> files)
        renamed_files_by_folder_and_marke[factorgroup][marke].append((file_path, filename))

    return renamed_files_by_folder_and_marke, file_to_factorgroup, marken_index


def generate_updated_pdf_using_helper(renamed_files_by_folder_and_marke: Dict, marken_index: Dict,
                                       output_path: Path) -> Tuple[bool, str]:
    """
    Generate PDF using THE EXACT SAME logic as helper.py's generate_filename_based_pdf_report_with_extensions
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib import colors
        from reportlab.platypus import Image as RLImage
        from pathlib import Path

        doc = SimpleDocTemplate(str(output_path), pagesize=A4, leftMargin=20, rightMargin=20,
                               topMargin=40, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()

        # Create brand columns - EXACT same as helper.py
        nummer_zu_marke = {v: k for k, v in marken_index.items()}
        marken_spalten = sorted(nummer_zu_marke.items())

        # Sort folders naturally by blocknummer - EXACT same as helper.py
        def get_folder_sort_key(folder_name):
            match = re.match(r'^(\d+)', folder_name)
            return int(match.group(1)) if match else 999

        sorted_folders = sorted(renamed_files_by_folder_and_marke.keys(), key=get_folder_sort_key)

        # Process each folder - EXACT same structure as helper.py
        for folder in sorted_folders:
            elements.append(Paragraph(f"<b>Folder: {folder}</b>", styles['Heading2']))

            abschnitt = renamed_files_by_folder_and_marke[folder]

            # Count UNIQUE base files per brand (considering deduplication) - EXACT same as helper.py
            brand_counts = {}
            for nummer, marke in marken_spalten:
                files = abschnitt.get(marke, [])
                # Group by base filename (without extension) to count unique files
                unique_bases = {}
                for pfad, original_name in files:
                    base_name = Path(original_name).stem
                    if base_name not in unique_bases:
                        unique_bases[base_name] = []
                    unique_bases[base_name].append(Path(original_name).suffix.lower())
                brand_counts[marke] = len(unique_bases)

            total = sum(brand_counts.values())
            lines = []
            for nummer, marke in marken_spalten:
                count = brand_counts.get(marke, 0)
                lines.append(f"{nummer} ({marke}): {count}")
            lines.append(f"Total: {total}")
            elements.append(Paragraph("<br/>".join(lines), styles['Normal']))
            elements.append(Spacer(1, 10))

            # Create table - EXACT same as helper.py
            headers = [f"{nummer} ({marke})" for nummer, marke in marken_spalten]
            data = [headers]
            col_data = []
            max_rows = 0

            for _, marke in marken_spalten:
                eintraege = abschnitt.get(marke, [])
                zellen = []
                for pfad, original_name in eintraege:
                    # Use helper function to create cell
                    from helper import get_asset_cell
                    cell = get_asset_cell(pfad, original_name, len(headers))
                    zellen.append(cell)
                col_data.append(zellen)
                max_rows = max(max_rows, len(zellen))

            for i in range(max_rows):
                row = []
                for col in col_data:
                    row.append(col[i] if i < len(col) else "")
                data.append(row)

            col_width = (A4[0] - 40) / len(headers)
            t = Table(data, colWidths=[col_width] * len(headers))
            t.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            elements.append(t)
            elements.append(PageBreak())

        # Total overview - EXACT same as helper.py
        elements.append(Paragraph("<b>Total Overview</b>", styles['Heading2']))
        global_counts = defaultdict(int)
        for folder_data in renamed_files_by_folder_and_marke.values():
            for marke, daten in folder_data.items():
                global_counts[marke] += len(daten)

        gesamt = 0
        summary = []
        for nummer, marke in marken_spalten:
            count = global_counts[marke]
            summary.append(f"{marke}: {count} assets")
            gesamt += count
        summary.append(f"Total number of all assets: {gesamt}")
        elements.append(Paragraph("<br/>".join(summary), styles['Normal']))

        # Add brand pages - EXACT same as helper.py
        from helper import add_brand_pages_with_png_extensions
        add_brand_pages_with_png_extensions(elements, marken_spalten, renamed_files_by_folder_and_marke, marken_index)

        # Build PDF
        doc.build(elements)

        return True, "PDF generated successfully"

    except Exception as e:
        import traceback
        return False, f"Error generating PDF: {str(e)}\n{traceback.format_exc()}"


def generate_updated_excel_using_helper(processed_files_dir: Path, marken_index: Dict,
                                         file_to_factorgroup: Dict, png_to_ignore: Set[str],
                                         output_path: Path) -> Tuple[bool, str]:
    """
    Generate Excel using THE EXACT SAME logic as helper.py's generate_excel_report
    """
    try:
        from openpyxl.styles import PatternFill

        nummer_zu_marke = {v: k for k, v in marken_index.items()}
        data = []

        # Process all files - EXACT same logic as helper.py
        for file in sorted(processed_files_dir.iterdir()):
            if not file.is_file():
                continue

            # Skip ignored PNGs
            if file.name in png_to_ignore:
                continue

            name = file.stem
            file_ext = file.suffix.lower() if file.suffix else ''

            match = re.match(r"(\d{2})B(\d{2})([a-z0-9]+)", name, re.IGNORECASE)
            if match:
                markennummer, blocknummer, marke = match.groups()
                factor = nummer_zu_marke.get(markennummer, marke)

                factorgroup = file_to_factorgroup.get(file.name, f"{blocknummer}Unknown")
                factorgroup = factorgroup.replace('_', '')  # Remove all underscores - EXACT same as helper.py

                only_id = name
                is_b20_id = False
                is_non_image_file = False

                if "B20" in only_id.upper():
                    is_b20_id = True
                    language_value = only_id + file_ext
                # Highlight ALL non-image formats - EXACT same as helper.py
                elif file_ext in NON_IMAGE_EXTENSIONS:
                    is_non_image_file = True
                    if file_ext == '.txt':
                        try:
                            with open(file, 'r', encoding='utf-8', errors='ignore') as txt_file:
                                txt_content = txt_file.read().strip()
                                language_value = txt_content if txt_content else "[Empty file]"
                        except Exception as e:
                            language_value = f"[Error reading file: {str(e)}]"
                    else:
                        language_value = name + file_ext
                else:
                    language_value = name + file_ext

                data.append({
                    "factor": factor,
                    "factorgroup": factorgroup,
                    "ID": only_id,
                    "Language": language_value,
                    "highlight_yellow": is_non_image_file or is_b20_id
                })

        # Create Assets sheet - EXACT same as helper.py
        df_assets = pd.DataFrame(data, columns=["factor", "factorgroup", "ID", "Language"])

        # Create Reordered sheet - EXACT same logic as helper.py
        reordered_data = []
        for _, row in df_assets.iterrows():
            raw_factor = str(row["factorgroup"])
            group_prefix = raw_factor[:2]
            try:
                group = str(int(group_prefix))
            except ValueError:
                group = group_prefix

            clean_factor = str(row["factorgroup"]).replace('_', '')
            clean_factorgroup = str(row["factor"]).replace('_', '')

            reordered_data.append({
                "Group": group,
                "factor": clean_factor,
                "factorgroup": clean_factorgroup,
                "ID": row["ID"],
                "Language": row["Language"]
            })

        df_reordered = pd.DataFrame(reordered_data, columns=["Group", "factor", "factorgroup", "ID", "Language"])
        df_reordered['factor'] = df_reordered['factor'].astype(str)

        df_reordered['Group_num'] = pd.to_numeric(df_reordered['Group'], errors='coerce').fillna(999).astype(int)
        df_reordered['factor_num'] = df_reordered['factor'].str.extract(r'^(\d+)')[0]
        df_reordered['factor_num'] = pd.to_numeric(df_reordered['factor_num'], errors='coerce').fillna(999).astype(int)

        df_reordered = df_reordered.sort_values(
            by=['Group_num', 'factor_num', 'factor'],
            ascending=[True, True, True]
        ).reset_index(drop=True)

        df_reordered = df_reordered.drop(['factor_num', 'Group_num'], axis=1)

        # Write to Excel with highlighting - EXACT same as helper.py
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_assets.to_excel(writer, index=False, sheet_name="Assets")
            df_reordered.to_excel(writer, index=False, sheet_name="Reordered")

            workbook = writer.book
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            # Highlight Assets sheet
            assets_sheet = workbook['Assets']
            language_col_idx = 4
            for row_idx in range(2, len(df_assets) + 2):
                cell = assets_sheet.cell(row=row_idx, column=language_col_idx)
                if row_idx - 2 < len(data):
                    if data[row_idx - 2].get("highlight_yellow", False):
                        cell.fill = yellow_fill

            # Highlight Reordered sheet
            reordered_sheet = workbook['Reordered']
            language_col_idx = 5
            id_col_idx = 4

            id_to_highlight = {item["ID"]: item.get("highlight_yellow", False) for item in data}

            for row_idx in range(2, len(df_reordered) + 2):
                cell = reordered_sheet.cell(row=row_idx, column=language_col_idx)
                id_cell = reordered_sheet.cell(row=row_idx, column=id_col_idx)
                id_value = id_cell.value
                if id_value and id_to_highlight.get(id_value, False):
                    cell.fill = yellow_fill

        return True, "Excel generated successfully"

    except Exception as e:
        import traceback
        return False, f"Error generating Excel: {str(e)}\n{traceback.format_exc()}"


def updated_file_processor_tab():
    """
    Tab 6: Process updated zip files from Tab 5
    Uses EXACT same logic as helper.py to maintain structure
    """
    st.header("Rerun Matrix")



    # Initialize session state
    if 'tab6_processing_complete' not in st.session_state:
        st.session_state.tab6_processing_complete = False
    if 'tab6_data' not in st.session_state:
        st.session_state.tab6_data = {}

    uploaded_file = st.file_uploader(
        "Upload Modified ZIP File (from Tab 5)",
        type=['zip'],
        key='tab6_uploader'
    )

    # Reset processing state when new file is uploaded
    if uploaded_file and 'tab6_uploaded_file_name' in st.session_state:
        if st.session_state.tab6_uploaded_file_name != uploaded_file.name:
            st.session_state.tab6_processing_complete = False
            st.session_state.tab6_data = {}

    if uploaded_file:
        st.session_state.tab6_uploaded_file_name = uploaded_file.name

        if not st.session_state.tab6_processing_complete:
            with st.spinner("🔍 Extracting and analyzing uploaded zip file..."):
                try:
                    temp_dir = extract_uploaded_zip(uploaded_file)
                    structure = read_zip_structure(temp_dir)

                    st.session_state.tab6_data = {
                        'temp_dir': temp_dir,
                        'structure': structure
                    }
                    st.session_state.tab6_processing_complete = True
                    st.rerun()

                except Exception as e:
                    st.error(f"❌ Error processing zip file: {e}")
                    import traceback
                    st.code(traceback.format_exc())
                    return

        # Display extracted contents
        if st.session_state.tab6_processing_complete:
            structure = st.session_state.tab6_data['structure']

            if not structure['excel_data']:
                st.error("❌ Cannot proceed: Excel file is required to extract folder structure and metadata.")
                st.info("💡 Make sure your zip contains an Excel file starting with `Brand_Assets_Overview` in the `reports/` folder.")
            else:
                if st.button("🎯 Analyze Files & Generate Reports", type="primary"):
                    with st.spinner("Processing files and generating reports..."):
                        try:
                            # Step 1: Extract metadata from existing Excel
                            metadata = extract_metadata_from_excel(structure['excel_data'])

                            st.info(f"📊 Found {len(metadata['existing_files'])} existing files in {len(metadata['folder_names'])} folders")

                            # Step 2: Find PNG duplicates
                            png_to_ignore = find_duplicate_pngs(structure['processed_files'])

                            if png_to_ignore:
                                st.info(f"ℹ️ Ignoring {len(png_to_ignore)} PNG files that duplicate video/audio files")

                            # Step 3: Analyze all processed files using helper.py logic
                            renamed_files_by_folder, file_to_factorgroup, marken_index = analyze_processed_files(
                                structure['processed_files'],
                                metadata,
                                png_to_ignore
                            )

                            # Step 4: Generate updated PDF using EXACT helper.py logic
                            temp_dir_path = Path(st.session_state.tab6_data['temp_dir'])
                            pdf_output_path = temp_dir_path / "Brand_Assets_Report_Updated.pdf"

                            pdf_success, pdf_msg = generate_updated_pdf_using_helper(
                                renamed_files_by_folder,
                                marken_index,
                                pdf_output_path
                            )

                            if pdf_success:
                                st.success(f"✅ PDF: {pdf_msg}")
                            else:
                                st.error(f"❌ PDF: {pdf_msg}")

                            # Step 5: Generate updated Excel using EXACT helper.py logic
                            excel_output_path = temp_dir_path / "IcAt_Overview_Final_Updated.xlsx"

                            excel_success, excel_msg = generate_updated_excel_using_helper(
                                structure['processed_files_dir'],
                                marken_index,
                                file_to_factorgroup,
                                png_to_ignore,
                                excel_output_path
                            )

                            if excel_success:
                                st.success(f"✅ Excel: {excel_msg}")
                            else:
                                st.error(f"❌ Excel: {excel_msg}")

                            # Store results
                            st.session_state.tab6_data['pdf_updated'] = pdf_output_path if pdf_success else None
                            st.session_state.tab6_data['excel_updated'] = excel_output_path if excel_success else None
                            st.session_state.tab6_data['png_ignored'] = png_to_ignore
                            st.session_state.tab6_data['reports_generated'] = True
                            st.session_state.tab6_data['metadata'] = metadata

                            st.rerun()

                        except Exception as e:
                            st.error(f"❌ Error during processing: {str(e)}")
                            import traceback
                            st.code(traceback.format_exc())

            # Show download buttons if reports were generated
            if st.session_state.tab6_data.get('reports_generated'):
                st.success("Processing complete!")

                # Show processing info
                if st.session_state.tab6_data.get('metadata'):
                    metadata = st.session_state.tab6_data['metadata']
                    with st.expander("ℹ️ Processing Details"):
                        st.write(f"**Folder names detected:** {len(metadata['folder_names'])}")
                        for folder in sorted(metadata['folder_names']):
                            st.text(f"  • {folder}")

                        st.write(f"\n**Brand index:**")
                        for brand, num in sorted(metadata['marken_index'].items()):
                            st.text(f"  • {num}: {brand}")

                # Show ignored files info
                if st.session_state.tab6_data.get('png_ignored'):
                    with st.expander("⚠️ Ignored PNG Files (duplicates of video/audio)"):
                        for png_file in sorted(st.session_state.tab6_data['png_ignored']):
                            st.text(f"  • {png_file}")

                # Option to start over
                if st.button("🔄 Process New File", help="Clear results and upload a new file"):
                    if 'temp_dir' in st.session_state.tab6_data:
                        try:
                            shutil.rmtree(st.session_state.tab6_data['temp_dir'])
                        except:
                            pass
                    st.session_state.tab6_processing_complete = False
                    st.session_state.tab6_data = {}
                    st.rerun()

                st.markdown("---")

                # Combined download option (recommended)
                st.subheader("📦 Complete Package Download")
                st.markdown("**Recommended:** Download everything in one convenient package")

                # Create combined zip with updated files
                if st.session_state.tab6_data.get('pdf_updated') and st.session_state.tab6_data.get('excel_updated'):
                    import io
                    combined_zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(combined_zip_buffer, 'w', zipfile.ZIP_DEFLATED) as combined_zip:
                        # Add processed files
                        for file in structure['processed_files']:
                            if file.is_file():
                                combined_zip.write(file, f"processed_files/{file.name}")

                        # Add updated PDF report
                        pdf_path = st.session_state.tab6_data['pdf_updated']
                        if pdf_path.exists():
                            with open(pdf_path, 'rb') as f:
                                combined_zip.writestr("reports/Brand_Assets_Report.pdf", f.read())

                        # Add updated Excel report
                        excel_path = st.session_state.tab6_data['excel_updated']
                        if excel_path.exists():
                            with open(excel_path, 'rb') as f:
                                combined_zip.writestr("reports/Brand_Assets_Overview.xlsx", f.read())

                    combined_zip_buffer.seek(0)

                    st.download_button(
                        label="🎁 Download Complete Package (All Files + Reports)",
                        data=combined_zip_buffer.getvalue(),
                        file_name="Brand_Assets_Complete_Package_Updated.zip",
                        mime="application/zip",
                        type="primary"
                    )

                st.markdown("---")

                # Individual download options
                st.subheader("📁 Individual Downloads")
                st.markdown("Or download items separately:")

                col1, col2, col3 = st.columns(3)

                with col1:
                    if st.session_state.tab6_data.get('pdf_updated'):
                        pdf_path = st.session_state.tab6_data['pdf_updated']
                        if pdf_path.exists():
                            with open(pdf_path, 'rb') as f:
                                st.download_button(
                                    label="📄 PDF Report",
                                    data=f.read(),
                                    file_name="Brand_Assets_Report.pdf",
                                    mime="application/pdf"
                                )
                    else:
                        st.info("PDF report not available")

                with col2:
                    if st.session_state.tab6_data.get('excel_updated'):
                        excel_path = st.session_state.tab6_data['excel_updated']
                        if excel_path.exists():
                            with open(excel_path, 'rb') as f:
                                st.download_button(
                                    label="📊 Excel Report",
                                    data=f.read(),
                                    file_name="Brand_Assets_Overview.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                    else:
                        st.info("Excel report not available")

                with col3:
                    if structure['processed_files_dir']:
                        # Create zip of processed files
                        import io
                        zip_buffer = io.BytesIO()
                        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                            for file in structure['processed_files']:
                                if file.is_file():
                                    zip_file.write(file, file.name)
                        zip_buffer.seek(0)

                        st.download_button(
                            label="📦 Processed Files",
                            data=zip_buffer.getvalue(),
                            file_name="Brand_Assets_Processed.zip",
                            mime="application/zip"
                        )
                    else:
                        st.info("Processed files not available")

                st.markdown("---")
                st.info("💡 **Tip:** Downloads will remain available until you upload a new file or click 'Process New File'")

    else:
        st.info("👆 Please upload a zip file to begin.")

        with st.expander("📖 How to Use This Tool"):
            st.markdown("""
            ### Workflow:

            1. **Get zip from Tab 5**: Download "Complete Package" from Tab 5

            2. **Optional: Update files manually**:
               - Extract the zip locally
               - Add/remove/modify files in `processed_files/` folder
               - Keep the filename structure: `01B11brandname01description.ext`
               - Re-zip maintaining folder structure

            3. **Upload here**: Upload the zip file

            4. **Generate reports**: Click "Analyze Files & Generate Reports"
            """)


if __name__ == "__main__":
    updated_file_processor_tab()
