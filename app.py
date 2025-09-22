import io
import os
import re
import shutil
import struct
import tempfile
import zipfile
from collections import defaultdict
from pathlib import Path
import cv2
import numpy as np
import pandas as pd
import streamlit as st
import xlrd
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Image as RLImage, 
    Paragraph, Spacer, PageBreak, KeepInFrame
)
import xml.etree.ElementTree as ET
import openpyxl
# Configure Streamlit page
st.set_page_config(
    page_title="Brand Asset Management Tools​",
    page_icon="🔧",
    layout="wide"
)

# Defined file types
ALLOWED_IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.webp'}
ALLOWED_TEXT_EXTENSIONS = {'.txt', '.md', '.csv'}
ALLOWED_VIDEO_EXTENSIONS = {'.mp4', '.avi', '.mov', '.wmv', '.flv', '.mkv', '.webm', '.m4v', '.3gp', '.ogv'}
ALLOWED_AUDIO_EXTENSIONS = {'.mp3', '.wav', '.aac', '.flac', '.ogg', '.m4a', '.wma'}
ALL_ALLOWED_EXTENSIONS = ALLOWED_IMAGE_EXTENSIONS.union(ALLOWED_TEXT_EXTENSIONS).union(ALLOWED_VIDEO_EXTENSIONS).union(ALLOWED_AUDIO_EXTENSIONS)


# Helper functions from both files
def extract_zip_to_temp(uploaded_file):
    """Extract uploaded zip file to temporary directory"""
    temp_dir = tempfile.mkdtemp()
    with zipfile.ZipFile(uploaded_file, 'r') as zip_ref:
        zip_ref.extractall(temp_dir)
    return temp_dir

def extract_images_from_xlsx(xlsx_file, output_dir):
    """Extract images from Excel file to specified directory"""
    count = 0
    extracted_files = []
    
    try:
        with zipfile.ZipFile(xlsx_file, 'r') as z:
            for file_info in z.infolist():
                if file_info.filename.startswith("xl/media/"):
                    filename = Path(file_info.filename).name
                    target_path = Path(output_dir) / filename
                    
                    with z.open(file_info) as source, open(target_path, 'wb') as target:
                        shutil.copyfileobj(source, target)
                    
                    extracted_files.append(str(target_path))
                    count += 1
        
        return count, extracted_files, None
    except Exception as e:
        return 0, [], str(e)

def extract_brand(file_stem):
    """Extract brand name from filename using pattern matching"""
    parts = re.split(r'[ _\-]', file_stem)
    return re.sub(r'[^a-z0-9]', '', parts[0].lower()) if parts else "unknown"

def erkenne_marken_aus_ordnern(input_folder):
    """Recognize brands from subfolders (original method)"""
    try:
        return sorted([
            f for f in os.listdir(input_folder)
            if os.path.isdir(os.path.join(input_folder, f)) and not f.startswith('.')
        ])
    except Exception as e:
        st.error(f"Error reading folders: {str(e)}")
        return []

def get_files_by_marke(input_folder, marken):
    """Get files per brand from folders"""
    spalteninhalte = {marke: [] for marke in marken}
    
    for marke in marken:
        folder_path = os.path.join(input_folder, marke)
        if os.path.isdir(folder_path):
            try:
                dateien = [
                    os.path.join(folder_path, f)
                    for f in os.listdir(folder_path)
                    if not f.startswith(".") and os.path.splitext(f)[1].lower() in ALL_ALLOWED_EXTENSIONS
                ]
                spalteninhalte[marke] = dateien
            except Exception as e:
                st.warning(f"Error reading files from {marke}: {str(e)}")
                spalteninhalte[marke] = []
    
    return spalteninhalte

def get_asset_cell(file_path, filename, col_count):
    """Create optimized asset cell for PDF with proper transparency handling"""
    ext = Path(file_path).suffix.lower()
    styles = getSampleStyleSheet()
    
    try:
        if ext in ALLOWED_IMAGE_EXTENSIONS:
            with PILImage.open(file_path) as img:
                # Preserve transparency if available
                if img.mode in ('RGBA', 'LA') or (img.mode == 'P' and 'transparency' in img.info):
                    # Create a white background for the PDF (transparency doesn't work well in PDF)
                    background = PILImage.new('RGB', img.size, (255, 255, 255))
                    if img.mode == 'RGBA':
                        # Paste the image onto the white background using the alpha channel as mask
                        background.paste(img, mask=img.split()[-1])
                    else:
                        # For other transparent modes, convert to RGBA first
                        img_rgba = img.convert('RGBA')
                        background.paste(img_rgba, mask=img_rgba.split()[-1])
                    img = background.convert('RGB')
                else:
                    # Convert non-transparent images to RGB
                    img = img.convert('RGB')
                
                orig_width, orig_height = img.size
                available_width = A4[0] - 40
                max_width = available_width / col_count
                max_height = 120

                aspect_ratio = orig_width / orig_height
                width = max_width
                height = width / aspect_ratio

                if height > max_height:
                    height = max_height
                    width = height * aspect_ratio

                buffer = io.BytesIO()
                img.save(buffer, format='PNG')
                buffer.seek(0)
                
                rl_img = RLImage(buffer, width=width, height=height)

                frame = KeepInFrame(max_width, max_height + 30, content=[
                    rl_img,
                    Paragraph(filename, styles['Normal'])
                ], hAlign='CENTER')

                return frame
        else:
            return Paragraph(f"{filename}", styles['Normal'])
    except Exception as e:
        return Paragraph(f"[Image Error]<br/>{filename}", styles['Normal'])
    

def build_marken_einzelseiten(alle_dateien_pro_marke, styles, gesamtbreite):
    """Build individual brand overview pages (original method)"""
    elements = []
    max_bildhoehe = 100

    for marke, dateien in alle_dateien_pro_marke.items():
        elements.append(Paragraph(f"Brand Overview: {marke}", styles['Heading2']))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"<b>Total Assets:</b> {len(dateien)}", styles['Normal']))
        elements.append(Spacer(1, 12))

        if not dateien:
            elements.append(Paragraph("No assets found in this brand folder.", styles['Normal']))
           # elements.append(PageBreak())
            continue

        table_data = []
        row_data = []
        spaltenbreite = gesamtbreite / 3

        for i, file_path in enumerate(dateien):
            ext = os.path.splitext(file_path)[1].lower()
            filename = os.path.basename(file_path)

            if ext in ALLOWED_IMAGE_EXTENSIONS:
                try:
                    pil_img = PILImage.open(file_path)
                    width, height = pil_img.size
                    aspect = height / width
                    new_width = spaltenbreite * 0.9
                    new_height = min(new_width * aspect, max_bildhoehe)
                    new_width = new_height / aspect

                    img = RLImage(file_path, width=new_width, height=new_height)
                    cell = [img, Spacer(1, 4), Paragraph(filename, styles['Normal'])]
                except Exception as e:
                    cell = Paragraph(f"{filename} (Error loading image)", styles['Normal'])
            else:
                cell = Paragraph(f"{filename}", styles['Normal'])

            row_data.append(cell)
            if len(row_data) == 3:
                table_data.append(row_data)
                row_data = []

        if row_data:
            while len(row_data) < 3:
                row_data.append("")
            table_data.append(row_data)

        if table_data:
            table = Table(table_data, colWidths=[spaltenbreite] * 3)
            table.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))
            elements.append(table)
        
        elements.append(PageBreak())
    
    return elements

def add_brand_pages(elements, marken_spalten, renamed_files_by_folder_and_marke):
    """Add brand overview pages with optimized layout"""
    styles = getSampleStyleSheet()
    for nummer, marke in marken_spalten:
        assets = []
        for folder in renamed_files_by_folder_and_marke:
            assets.extend(renamed_files_by_folder_and_marke[folder].get(marke, []))

        if not assets:
            continue

        elements.append(PageBreak())
        elements.append(Paragraph(f"<b>Brand Overview: {nummer} – {marke}</b>", styles['Heading2']))
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(f"Number of assets: {len(assets)}", styles['Normal']))
        elements.append(Spacer(1, 10))

        headers = ["Asset"] * 4
        data = [headers]
        row = []
        for i, (pfad, name) in enumerate(assets):
            cell = get_asset_cell(pfad, name, 4)
            row.append(cell)
            if len(row) == 4:
                data.append(row)
                row = []
        if row:
            row.extend([""] * (4 - len(row)))
            data.append(row)

        table = Table(data, colWidths=(A4[0] - 40) / 4)
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(table)

def analyze_files_by_filename(input_folder):
    """Analyze files by extracting brand names from filenames"""
    dateien = []
    marken_set = set()
    renamed_files_by_folder_and_marke = defaultdict(lambda: defaultdict(list))

    for subfolder in sorted(Path(input_folder).iterdir()):
        if not subfolder.is_dir():
            continue

        for file in sorted(subfolder.iterdir()):
            if file.suffix.lower() not in ALL_ALLOWED_EXTENSIONS:
                continue
            marke = extract_brand(file.stem)
            marken_set.add(marke)
            dateien.append({
                "original_file": file,
                "original_folder": subfolder.name,
                "marke": marke
            })

    for eintrag in dateien:
        orig = eintrag["original_file"]
        marke = eintrag["marke"]
        folder_name = eintrag["original_folder"]
        renamed_files_by_folder_and_marke[folder_name][marke].append((orig, orig.name))

    return marken_set, renamed_files_by_folder_and_marke, dateien

def generate_pdf_report(input_folder, erste_marke=None):
    """Generate the complete PDF report (original method)"""
    styles = getSampleStyleSheet()
    gesamtbreite = 480

    marken = erkenne_marken_aus_ordnern(input_folder)
    if not marken:
        return None, "No brand folders found in the uploaded zip file."

    # Reorder brands if specified
    if erste_marke and erste_marke in marken:
        marken = [erste_marke] + [m for m in marken if m != erste_marke]

    dateien_pro_marke = get_files_by_marke(input_folder, marken)
    total_all = sum(len(dateien_pro_marke[marke]) for marke in marken)

    alle_elements = []

    # Overview page
    alle_elements.append(Paragraph("Brand Assets Overview", styles['Title']))
    alle_elements.append(Spacer(1, 20))
    
    übersicht_text = ", ".join([f"{marke}: {len(dateien_pro_marke[marke])}" for marke in marken])
    alle_elements.append(Paragraph(f"<b>Assets per brand:</b> {übersicht_text}", styles['Normal']))
    alle_elements.append(Spacer(1, 12))
    alle_elements.append(Paragraph(f"<b>Total assets:</b> {total_all}", styles['Normal']))
    alle_elements.append(PageBreak())

    # Individual brand pages
    markenseiten = build_marken_einzelseiten(dateien_pro_marke, styles, gesamtbreite)
    alle_elements.extend(markenseiten)

    # Generate PDF
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
    
    try:
        doc.build(alle_elements)
        pdf_buffer.seek(0)
        return pdf_buffer, None
    except Exception as e:
        return None, f"Error generating PDF: {str(e)}"

def generate_filename_based_pdf_report(input_folder, erste_marke=None, processed_files_mapping=None):
    """Generate PDF report based on filename brand analysis with final processed labels"""
    marken_set, renamed_files_by_folder_and_marke, all_files = analyze_files_by_filename(input_folder)
    
    if not marken_set:
        return None, "No brands found in filenames."

    # Create brand index
    marken_index = {}
    if erste_marke and erste_marke in marken_set:
        marken_index[erste_marke] = "01"
        aktuelle_nummer = 2
        for marke in sorted(marken_set):
            if marke != erste_marke:
                marken_index[marke] = f"{aktuelle_nummer:02d}"
                aktuelle_nummer += 1
    else:
        for i, marke in enumerate(sorted(marken_set), 1):
            marken_index[marke] = f"{i:02d}"

    # Generate PDF
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, leftMargin=20, rightMargin=20, topMargin=40, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    nummer_zu_marke = {v: k for k, v in marken_index.items()}
    marken_spalten = sorted(nummer_zu_marke.items())

    # Process each folder with final processed names
    for folder in sorted(renamed_files_by_folder_and_marke):
        elements.append(Paragraph(f"<b>Folder: {folder}</b>", styles['Heading2']))

        abschnitt = renamed_files_by_folder_and_marke[folder]
        total = 0
        lines = []
        for nummer, marke in marken_spalten:
            count = len(abschnitt.get(marke, []))
            total += count
            lines.append(f"{nummer} ({marke}): {count}")
        lines.append(f"Total: {total}")
        elements.append(Paragraph("<br/>".join(lines), styles['Normal']))
        elements.append(Spacer(1, 10))

        headers = [f"{nummer} ({marke})" for nummer, marke in marken_spalten]
        data = [headers]
        col_data = []
        max_rows = 0
        
        for _, marke in marken_spalten:
            eintraege = abschnitt.get(marke, [])
            zellen = []
            for original_path, original_name in eintraege:
                # Generate the final processed ID name
                blocknummer = re.sub(r'\D', '', folder)[:2].zfill(2)
                markennummer = marken_index[marke]
                cleaned = get_cleaned_filename_without_brand(original_name, marke)
                # Create the final ID without file extension for display
                final_id = f"{markennummer}B{blocknummer}{marke}{cleaned}"
                
                # Create cell with original path but final ID label
                cell = get_asset_cell(original_path, final_id, len(headers))
                zellen.append(cell)
            col_data.append(zellen)
            max_rows = max(max_rows, len(zellen))

        for i in range(max_rows):
            row = []
            for col in col_data:
                row.append(col[i] if i < len(col) else "")
            data.append(row)

        col_width = (A4[0] - 40) / len(headers)
        t = Table(data, colWidths=[col_width] * len(headers))
        t.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(t)
        elements.append(PageBreak())

    # Total overview
    elements.append(Paragraph("<b>Total Overview</b>", styles['Heading2']))
    global_counts = defaultdict(int)
    for folder_data in renamed_files_by_folder_and_marke.values():
        for marke, daten in folder_data.items():
            global_counts[marke] += len(daten)
    
    gesamt = 0
    summary = []
    for nummer, marke in marken_spalten:
        count = global_counts[marke]
        summary.append(f"{marke}: {count} assets")
        gesamt += count
    summary.append(f"Total number of all assets: {gesamt}")
    elements.append(Paragraph("<br/>".join(summary), styles['Normal']))

    # Add brand pages with final processed labels
    add_brand_pages_with_final_labels(elements, marken_spalten, renamed_files_by_folder_and_marke, marken_index)

    try:
        doc.build(elements)
        pdf_buffer.seek(0)
        return pdf_buffer, None
    except Exception as e:
        return None, f"Error generating PDF: {str(e)}"

def add_brand_pages_with_final_labels(elements, marken_spalten, renamed_files_by_folder_and_marke, marken_index):
    """Add brand overview pages with final processed ID labels"""
    styles = getSampleStyleSheet()
    for nummer, marke in marken_spalten:
        assets = []
        for folder in renamed_files_by_folder_and_marke:
            folder_assets = renamed_files_by_folder_and_marke[folder].get(marke, [])
            for original_path, original_name in folder_assets:
                # Generate the final processed ID name (same logic as in Excel generation)
                blocknummer = re.sub(r'\D', '', folder)[:2].zfill(2)
                markennummer = marken_index[marke]
                cleaned = get_cleaned_filename_without_brand(original_name, marke)
                final_id = f"{markennummer}B{blocknummer}{marke}{cleaned}"
                assets.append((original_path, final_id))

        if not assets:
            continue

        elements.append(PageBreak())
        elements.append(Paragraph(f"<b>Brand Overview: {nummer} – {marke}</b>", styles['Heading2']))
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(f"Number of assets: {len(assets)}", styles['Normal']))
        elements.append(Spacer(1, 10))

        headers = ["Asset"] * 4
        data = [headers]
        row = []
        for i, (original_path, final_id) in enumerate(assets):
            # Use final processed ID as the label
            cell = get_asset_cell(original_path, final_id, 4)
            row.append(cell)
            if len(row) == 4:
                data.append(row)
                row = []
        if row:
            row.extend([""] * (4 - len(row)))
            data.append(row)

        table = Table(data, colWidths=(A4[0] - 40) / 4)
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(table)

def generate_two_section_pdf_report(input_folder, erste_marke=None):
    """Generate PDF report with two sections: by block and by brand - WITH EXTENSIONS"""
    marken_set, renamed_files_by_folder_and_marke, all_files = analyze_files_by_filename(input_folder)
    
    if not marken_set:
        return None, "No brands found in filenames."

    # Create brand index
    marken_index = {}
    if erste_marke and erste_marke in marken_set:
        marken_index[erste_marke] = "01"
        aktuelle_nummer = 2
        for marke in sorted(marken_set):
            if marke != erste_marke:
                marken_index[marke] = f"{aktuelle_nummer:02d}"
                aktuelle_nummer += 1
    else:
        for i, marke in enumerate(sorted(marken_set), 1):
            marken_index[marke] = f"{i:02d}"

    # Generate PDF
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, leftMargin=20, rightMargin=20, topMargin=40, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    nummer_zu_marke = {v: k for k, v in marken_index.items()}
    marken_spalten = sorted(nummer_zu_marke.items())

    # SECTION 1: BY BLOCK/FOLDER
    elements.append(Paragraph("<b>Section 1: Assets by Block</b>", styles['Title']))
    elements.append(Spacer(1, 20))

    for folder in sorted(renamed_files_by_folder_and_marke):
        elements.append(Paragraph(f"<b>Folder: {folder}</b>", styles['Heading2']))

        abschnitt = renamed_files_by_folder_and_marke[folder]
        total = 0
        lines = []
        for nummer, marke in marken_spalten:
            count = len(abschnitt.get(marke, []))
            total += count
            lines.append(f"{nummer} ({marke}): {count}")
        lines.append(f"Total: {total}")
        elements.append(Paragraph("<br/>".join(lines), styles['Normal']))
        elements.append(Spacer(1, 10))

        headers = [f"{nummer} ({marke})" for nummer, marke in marken_spalten]
        data = [headers]
        col_data = []
        max_rows = 0
        
        for _, marke in marken_spalten:
            eintraege = abschnitt.get(marke, [])
            zellen = []
            for p, original_name in eintraege:
                # Generate the ID name - KEEP THE FILE EXTENSION
                blocknummer = re.sub(r'\D', '', folder)[:2].zfill(2)
                markennummer = marken_index[marke]
                cleaned = get_cleaned_filename_without_brand(original_name, marke)
                # Keep the original file extension
                file_extension = Path(original_name).suffix.lower()
                id_name = f"{markennummer}B{blocknummer}{marke}{cleaned}{file_extension}"
                
                # Pass the full filename with extension to get_asset_cell
                cell = get_asset_cell(p, id_name, len(headers))
                zellen.append(cell)
            col_data.append(zellen)
            max_rows = max(max_rows, len(zellen))

        for i in range(max_rows):
            row = []
            for col in col_data:
                row.append(col[i] if i < len(col) else "")
            data.append(row)

        col_width = (A4[0] - 40) / len(headers)
        t = Table(data, colWidths=[col_width] * len(headers))
        t.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(t)
        elements.append(PageBreak())

    # SECTION 2: BY BRAND
    elements.append(Paragraph("<b>Section 2: Assets by Brand</b>", styles['Title']))
    elements.append(Spacer(1, 20))

    # Total overview first
    elements.append(Paragraph("<b>Total Overview</b>", styles['Heading2']))
    global_counts = defaultdict(int)
    for folder_data in renamed_files_by_folder_and_marke.values():
        for marke, daten in folder_data.items():
            global_counts[marke] += len(daten)
    
    gesamt = 0
    summary = []
    for nummer, marke in marken_spalten:
        count = global_counts[marke]
        summary.append(f"{marke}: {count} assets")
        gesamt += count
    summary.append(f"Total number of all assets: {gesamt}")
    elements.append(Paragraph("<br/>".join(summary), styles['Normal']))
    elements.append(PageBreak())
    first_brand = True
    # Individual brand pages
    for nummer, marke in marken_spalten:
        assets = []
        for folder in renamed_files_by_folder_and_marke:
            folder_assets = renamed_files_by_folder_and_marke[folder].get(marke, [])
            brand_counter = 1  # Start counting from 1 for each brand in each folder
            for pfad, original_name in folder_assets:
                # Generate the ID name with sequential count - KEEP THE FILE EXTENSION
                blocknummer = re.sub(r'\D', '', folder)[:2].zfill(2)
                markennummer = marken_index[marke]
                count_str = f"{brand_counter:02d}"
                cleaned = get_cleaned_filename_without_brand(original_name, marke)
                # Keep the original file extension
                file_extension = Path(original_name).suffix.lower()
                id_name = f"{markennummer}B{blocknummer}{marke}{count_str}{cleaned}{file_extension}"
                assets.append((pfad, id_name))
                brand_counter += 1

        if not assets:
            continue



        # Inside the loop for each brand
        if not first_brand:
            elements.append(PageBreak())
        else:
            first_brand = False
        elements.append(Paragraph(f"<b>Brand Overview: {nummer} – {marke}</b>", styles['Heading2']))
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(f"Assets per Brand: {', '.join([f'{k}: {len([a for a in assets if extract_brand(Path(a[1]).stem) == k])}' for k in [marke]])}", styles['Normal']))
        elements.append(Paragraph(f"Total assets: {len(assets)}", styles['Normal']))
        elements.append(Spacer(1, 10))
        

        headers = ["Asset"] * 4
        data = [headers]
        row = []
        for i, (pfad, id_name) in enumerate(assets):
            # Pass the full filename with extension to get_asset_cell
            cell = get_asset_cell(pfad, id_name, 4)
            row.append(cell)
            if len(row) == 4:
                data.append(row)
                row = []
        if row:
            row.extend([""] * (4 - len(row)))
            data.append(row)

        table = Table(data, colWidths=(A4[0] - 40) / 4)
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(table)

    try:
        doc.build(elements)
        pdf_buffer.seek(0)
        return pdf_buffer, None
    except Exception as e:
        return None, f"Error generating PDF: {str(e)}"

def brand_renamer_tool():
    st.header("Advanced Brand File Processor")
    st.markdown("Automatically rename and organize brand assets with final names and outputs (pdf overview, excel output) for programming")
   
    # Initialize session state for persistent results
    if 'processing_complete' not in st.session_state:
        st.session_state.processing_complete = False
    if 'processed_data' not in st.session_state:
        st.session_state.processed_data = {}
   
    uploaded_file = st.file_uploader(
        "Choose a ZIP file",
        type=['zip'],
    )
   
    # Reset processing state when new file is uploaded
    if uploaded_file and 'uploaded_file_name' in st.session_state:
        if st.session_state.uploaded_file_name != uploaded_file.name:
            st.session_state.processing_complete = False
            st.session_state.processed_data = {}
   
    if uploaded_file:
        # Store uploaded file name to detect changes
        st.session_state.uploaded_file_name = uploaded_file.name
       
        if not st.session_state.processing_complete:
            with st.spinner("Extracting and analyzing files..."):
                temp_dir = extract_zip_to_temp(uploaded_file)
                input_folder = Path(temp_dir)
               
                items = os.listdir(temp_dir)
                if len(items) == 1 and os.path.isdir(os.path.join(temp_dir, items[0])):
                    input_folder = Path(temp_dir) / items[0]
               
                output_folder = Path(temp_dir) / "output"
                output_folder.mkdir(exist_ok=True)
               
                marken_set, _, _ = analyze_files_by_filename(input_folder)
               
                if not marken_set:
                    st.error("No brands found in the uploaded files.")
                    return
               
                st.success(f"Detected brands: {', '.join(sorted(marken_set))}")
               
                # Brand numbering selection
                st.subheader("Brand Numbering")
                erste_marke = st.selectbox(
                    "Which brand should be number 01?",
                    options=sorted(marken_set),
                    index=0
                )
               
                marken_index = {erste_marke: "01"}
                aktuelle_nummer = 2
                for marke in sorted(marken_set):
                    if marke != erste_marke:
                        marken_index[marke] = f"{aktuelle_nummer:02d}"
                        aktuelle_nummer += 1
               
                if st.button("Process Files and Generate Reports", type="primary"):
                    with st.spinner("Processing files..."):
                        # Process and rename files
                        renamed_files, file_to_factorgroup = process_files(
                            input_folder,
                            marken_index,
                            output_folder
                        )
                       
                        # Generate PDF with TWO SECTIONS (by block and by brand)
                        pdf_buffer, error = generate_two_section_pdf_report(
                            input_folder,
                            erste_marke
                        )
                        if error:
                            st.error(f"PDF generation failed: {error}")
                       
                        excel_path = generate_excel_report(output_folder, marken_index, file_to_factorgroup)
                       
                        # Create individual file buffers
                        zip_buffer = io.BytesIO()
                        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                            for file in output_folder.iterdir():
                                zip_file.write(file, file.name)
                        zip_buffer.seek(0)
                       
                        # Create combined download with all files
                        combined_zip_buffer = io.BytesIO()
                        with zipfile.ZipFile(combined_zip_buffer, 'w', zipfile.ZIP_DEFLATED) as combined_zip:
                            # Add processed files
                            for file in output_folder.iterdir():
                                combined_zip.write(file, f"processed_files/{file.name}")
                           
                            # Add PDF report if available
                            if pdf_buffer:
                                combined_zip.writestr("reports/Brand_Assets_Report_Two_Sections.pdf", pdf_buffer.getvalue())
                           
                            # Add Excel report
                            with open(excel_path, 'rb') as excel_file:
                                combined_zip.writestr("reports/Brand_Assets_Overview.xlsx", excel_file.read())
                       
                        combined_zip_buffer.seek(0)
                       
                        # Store processed data in session state
                        st.session_state.processed_data = {
                            'pdf_buffer': pdf_buffer,
                            'excel_path': excel_path,
                            'zip_buffer': zip_buffer,
                            'combined_zip_buffer': combined_zip_buffer,
                            'temp_dir': temp_dir,
                            'renamed_files': renamed_files
                        }
                        st.session_state.processing_complete = True
                       
                        # Force rerun to show download buttons
                        st.rerun()
       
        # Show download options if processing is complete
        if st.session_state.processing_complete:
            st.success("✅ Processing complete! PDF now includes both sections: by block and by brand.")
           
            # Display a preview of processed names
            with st.expander("Preview Final Processed Labels"):
                if 'renamed_files' in st.session_state.processed_data:
                    st.write("Sample of final processed labels (as used in both PDF sections and Excel):")
                    count = 0
                    for folder_data in st.session_state.processed_data['renamed_files'].values():
                        for brand_data in folder_data.values():
                            for _, processed_name in brand_data:
                                if count < 10:  # Show first 10
                                    # Remove file extension for display
                                    label_without_ext = Path(processed_name).stem
                                    st.write(f"📄 {label_without_ext}")
                                    count += 1
                    if count >= 10:
                        st.info("Showing first 10 labels...")
           
            # Option to start over
            if st.button("🔄 Process New File", help="Clear results and upload a new file"):
                st.session_state.processing_complete = False
                st.session_state.processed_data = {}
                if 'temp_dir' in st.session_state.processed_data:
                    try:
                        shutil.rmtree(st.session_state.processed_data['temp_dir'])
                    except:
                        pass
                st.rerun()
           
            st.markdown("---")
           
            # Combined download option (recommended)
            st.subheader("📦 Complete Package Download")
            st.markdown("**Recommended:** Download everything in one convenient package")
           
            if st.session_state.processed_data.get('combined_zip_buffer'):
                st.download_button(
                    label="🎁 Download Complete Package (All Files + Reports)",
                    data=st.session_state.processed_data['combined_zip_buffer'].getvalue(),
                    file_name="Brand_Assets_Complete_Package.zip",
                    mime="application/zip",
                    type="primary"
                )
           
            st.markdown("---")
           
            # Individual download options
            st.subheader("📁 Individual Downloads")
            st.markdown("Or download items separately:")
           
            col1, col2, col3 = st.columns(3)
           
            with col1:
                if st.session_state.processed_data.get('pdf_buffer'):
                    st.download_button(
                        label="📄 PDF Report (Two Sections)",
                        data=st.session_state.processed_data['pdf_buffer'].getvalue(),
                        file_name="Brand_Assets_Report_Two_Sections.pdf",
                        mime="application/pdf"
                    )
                else:
                    st.info("PDF report not available")
           
            with col2:
                if st.session_state.processed_data.get('excel_path'):
                    try:
                        with open(st.session_state.processed_data['excel_path'], 'rb') as excel_file:
                            st.download_button(
                                label="📊 Excel Report (Final Labels)",
                                data=excel_file.read(),
                                file_name="Brand_Assets_Overview.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    except:
                        st.info("Excel report not available")
           
            with col3:
                if st.session_state.processed_data.get('zip_buffer'):
                    st.download_button(
                        label="📦 Processed Files",
                        data=st.session_state.processed_data['zip_buffer'].getvalue(),
                        file_name="Brand_Assets_Processed.zip",
                        mime="application/zip"
                    )
           
            st.markdown("---")
            st.success("💡 **Updated:** PDF now contains two distinct sections matching your requirements!")
   
    else:
        # Reset session state when no file is uploaded
        if st.session_state.processing_complete:
            st.session_state.processing_complete = False
            st.session_state.processed_data = {}
       
        st.info("👆 Please upload a zip file to get started.")
       
        with st.expander("📖 Instructions"):
            st.markdown("""
            """)

def pdf_generator_tool():
    st.header("Asset Overview​")
    st.markdown("Upload a zip file containing brand assets to generate asset overview by brand​")
    
    # Analysis method selection
    analysis_method = st.radio(
        "Choose:",
        ["Overview by Brand", "Overview by Asset Type"],
        help="Overview by Brand: assets are organized by brand folders (they do NOT include brandname_ (underscore \"_\") in filename) ​. Overview by Asset Type: assets are organized by blocks AND include brandname_ (underscore \"_\") in filename)​"
    )
    
    # File upload
    uploaded_file = st.file_uploader(
        "Choose a ZIP file",
        type=['zip'],
        key="pdf_zip_upload"
    )
    
    if uploaded_file is not None:
        try:
            # Extract zip file
            with st.spinner("Extracting zip file..."):
                temp_dir = extract_zip_to_temp(uploaded_file)
            
            # Find the actual input folder (in case zip has a root folder)
            input_folder = temp_dir
            items = os.listdir(temp_dir)
            if len(items) == 1 and os.path.isdir(os.path.join(temp_dir, items[0])):
                input_folder = os.path.join(temp_dir, items[0])
            
            if analysis_method == "Overview by Brand":
                # Original folder-based analysis
                marken = erkenne_marken_aus_ordnern(input_folder)
                
                if not marken:
                    st.error("No brand folders found in the uploaded zip file.")
                    st.info("Make sure your zip file contains folders named after your brands, with assets inside each folder.")
                    return
                
                st.success(f"Found brands: {', '.join(marken)}")
                
                # Get file counts for processing
                dateien_pro_marke = get_files_by_marke(input_folder, marken)
                
                # Options
                erste_marke = st.selectbox(
                    "Which brand should appear first in the overview?",
                    options=["Alphabetical order"] + marken
                )
                erste_marke = erste_marke if erste_marke != "Alphabetical order" else None
                
                # Generate PDF button
                if st.button("Generate PDF Report", type="primary"):
                    with st.spinner("Generating PDF report..."):
                        pdf_buffer, error = generate_pdf_report(input_folder, erste_marke)
                    
                    if error:
                        st.error(f"{error}")
                    elif pdf_buffer:
                        st.success("PDF report generated successfully!")
                        
                        # Download button
                        st.download_button(
                            label="📥 Download PDF Report",
                            data=pdf_buffer.getvalue(),
                            file_name="Brand_Assets_Report.pdf",
                            mime="application/pdf"
                        )
                        
            else:
                # Filename-based analysis (Overview by Asset Type)
                marken_set, renamed_files_by_folder_and_marke, all_files = analyze_files_by_filename(input_folder)
                
                if not marken_set:
                    st.error("No brands found in filenames.")
                    st.info("Make sure your files are named with brand identifiers at the beginning (e.g., 'Brand1_asset.jpg').")
                    return
                
                st.success(f"Found brands from filenames: {', '.join(sorted(marken_set))}")
                
                # Show analysis preview
                with st.expander("Analysis Preview"):
                    total_files = len(all_files)
                    for folder, brands in renamed_files_by_folder_and_marke.items():
                        st.write(f"**{folder}:**")
                        for brand, files in brands.items():
                            st.write(f"- {brand}: {len(files)} files")
                
                # Options
                erste_marke = st.selectbox(
                    "Which brand should be numbered as 01?",
                    options=["Alphabetical order"] + sorted(list(marken_set))
                )
                erste_marke = erste_marke if erste_marke != "Alphabetical order" else None
                
                # Generate PDF button
                if st.button("Generate PDF Report", type="primary"):
                    with st.spinner("Generating PDF report..."):
                        # Use the two-section report function for filename-based analysis
                        pdf_buffer, error = generate_two_section_pdf_report(input_folder, erste_marke)
                    
                    if error:
                        st.error(f"{error}")
                    elif pdf_buffer:
                        st.success("PDF report generated successfully! (Contains two sections: by Brand and by Asset Type)")
                        
                        # Download button
                        st.download_button(
                            label="📥 Download PDF Report (Two Sections)",
                            data=pdf_buffer.getvalue(),
                            file_name="IcAt_Overview by Brand and Asset Type.pdf",
                            mime="application/pdf"
                        )
        
        except zipfile.BadZipFile:
            st.error("Invalid zip file. Please upload a valid zip file.")
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
    
    else:
        st.info("👆 Please upload a zip file to get started.")
        
        # Instructions
        with st.expander("📖 Instructions"):
            st.markdown("""
            **How to use this tool:**
            
            **Overview by Brand:**
            - Create folders named after your brands
            - Place brand assets inside each brand folder
            - Assets do NOT need brand names in their filenames
            - Zip the entire structure
            - Generates single-section PDF report organized by brand
            
            **Overview by Asset Type:**
            - Name your files with brand identifiers at the beginning followed by an underscore (e.g., 'Brand1_asset.jpg', 'Brand2_document.pdf')
            - Organize files in any folder structure (blocks/categories)
            - The tool will extract brand names from filenames automatically
            - Generates two-section PDF report: Section 1 (by Asset Type/Block) and Section 2 (by Brand)
            
            **Supported file types:**
            - **Images:** JPG, JPEG, PNG, BMP, GIF, TIFF, WEBP
            - **Text files:** TXT, MD, CSV
            """)



def extract_images_from_xlsx_worksheet_specific(excel_file_path, output_dir, target_worksheet=None):
    """
    Extract images from specific worksheet in .xlsx files
    """
    extracted_files = []
    available_worksheets = []
    
    try:
        # First, get worksheet information using openpyxl
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_file_path, read_only=True)
            available_worksheets = wb.sheetnames
            wb.close()
        except ImportError:
            # Fallback: read worksheet names from zip structure
            with zipfile.ZipFile(excel_file_path, 'r') as z:
                try:
                    workbook_xml = z.read('xl/workbook.xml').decode('utf-8')
                    root = ET.fromstring(workbook_xml)
                    sheets = root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet')
                    available_worksheets = [sheet.get('name') for sheet in sheets]
                except:
                    available_worksheets = ["Sheet1"]  # Default fallback

        with zipfile.ZipFile(excel_file_path, 'r') as z:
            # Get all relationships and media files
            worksheet_media_map = {}
            
            # Parse worksheet relationships to find which images belong to which worksheet
            if target_worksheet and target_worksheet != "All worksheets":
                # Find the worksheet ID for the target worksheet
                target_sheet_id = None
                try:
                    workbook_xml = z.read('xl/workbook.xml').decode('utf-8')
                    root = ET.fromstring(workbook_xml)
                    sheets = root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet')
                    
                    for sheet in sheets:
                        if sheet.get('name') == target_worksheet:
                            target_sheet_id = sheet.get('sheetId')
                            break
                except Exception as e:
                    print(f"Error parsing workbook.xml: {e}")
                    target_sheet_id = "1"  # Default to first sheet

                if target_sheet_id:
                    # Parse worksheet relationships
                    try:
                        worksheet_rels_path = f'xl/worksheets/_rels/sheet{target_sheet_id}.xml.rels'
                        if worksheet_rels_path not in z.namelist():
                            # Try alternative naming
                            worksheet_rels_path = f'xl/worksheets/_rels/sheet1.xml.rels'
                        
                        if worksheet_rels_path in z.namelist():
                            rels_xml = z.read(worksheet_rels_path).decode('utf-8')
                            rels_root = ET.fromstring(rels_xml)
                            
                            # Find image relationships
                            relationships = rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship')
                            worksheet_images = []
                            
                            for rel in relationships:
                                target_path = rel.get('Target')
                                rel_type = rel.get('Type')
                                
                                # Check if this is an image relationship
                                if 'image' in rel_type.lower() and target_path:
                                    # Convert relative path to absolute path in zip
                                    if target_path.startswith('../'):
                                        img_path = 'xl/' + target_path[3:]
                                    else:
                                        img_path = f'xl/worksheets/{target_path}'
                                    
                                    if img_path in z.namelist():
                                        worksheet_images.append(img_path)
                            
                            # Extract only images from this worksheet
                            count = 0
                            for img_path in worksheet_images:
                                filename = Path(img_path).name
                                target_path = Path(output_dir) / f"ws_{target_worksheet}_{filename}"
                                
                                with z.open(img_path) as source, open(target_path, 'wb') as target_file:
                                    shutil.copyfileobj(source, target_file)
                                
                                extracted_files.append(str(target_path))
                                count += 1
                            
                            return count, extracted_files, None, available_worksheets
                            
                    except Exception as e:
                        print(f"Error parsing worksheet relationships: {e}")
                        # Fall back to extracting all images
                        pass

            # If target_worksheet is "All worksheets" or parsing failed, extract all images
            media_files = [f for f in z.infolist() if f.filename.startswith("xl/media/")]
            count = 0
            
            for file_info in media_files:
                filename = Path(file_info.filename).name
                if target_worksheet == "All worksheets":
                    target_path = Path(output_dir) / filename
                else:
                    target_path = Path(output_dir) / f"all_{filename}"
                
                with z.open(file_info) as source, open(target_path, 'wb') as target_file:
                    shutil.copyfileobj(source, target_file)
                
                extracted_files.append(str(target_path))
                count += 1

            return count, extracted_files, None, available_worksheets

    except Exception as e:
        return 0, [], f"Error processing .xlsx file: {str(e)}", available_worksheets


def extract_images_from_xls_worksheet_specific(xls_file_path, output_dir, target_worksheet=None):
    """
    Extract images from .xls files with basic worksheet awareness
    Note: .xls format has limitations for precise worksheet-specific extraction
    """
    extracted_files = []
    available_worksheets = []
    
    try:
        # Get worksheet information
        try:
            import xlrd
            workbook = xlrd.open_workbook(xls_file_path, formatting_info=True)
            available_worksheets = workbook.sheet_names()
            
            if target_worksheet and target_worksheet != "All worksheets" and target_worksheet in available_worksheets:
                # For .xls, we can only provide basic filtering
                # The binary extraction method extracts all images, but we can at least validate the worksheet exists
                sheet_index = available_worksheets.index(target_worksheet)
                print(f"Extracting from worksheet '{target_worksheet}' (index: {sheet_index})")
        except ImportError:
            available_worksheets = ["All worksheets"]
        except Exception as e:
            print(f"Error reading .xls worksheets: {e}")
            available_worksheets = ["Sheet1"]

        # Use the comprehensive binary extraction method
        count, extracted_files, error = extract_ole_images_from_xls_comprehensive(xls_file_path, output_dir)
        
        # Rename files to indicate worksheet limitation
        if target_worksheet and target_worksheet != "All worksheets" and count > 0:
            renamed_files = []
            for i, file_path in enumerate(extracted_files):
                old_path = Path(file_path)
                new_path = old_path.parent / f"xls_{target_worksheet}_{old_path.name}"
                try:
                    os.rename(old_path, new_path)
                    renamed_files.append(str(new_path))
                except:
                    renamed_files.append(file_path)  # Keep original if rename fails
            extracted_files = renamed_files

        warning_msg = None
        if target_worksheet and target_worksheet != "All worksheets":
            warning_msg = f"Note: .xls format limitations - extracted images may be from multiple worksheets, not just '{target_worksheet}'"

        return count, extracted_files, warning_msg, available_worksheets

    except Exception as e:
        return 0, [], f"Error processing .xls file: {str(e)}", available_worksheets


def extract_ole_images_from_xls_comprehensive(xls_file_path, output_dir):
    """
    Comprehensive extraction of embedded images from .xls files
    """
    extracted_files = []
    count = 0
    
    try:
        # Read the entire file as binary
        with open(xls_file_path, 'rb') as f:
            data = f.read()

        # Enhanced image signatures
        image_signatures = [
            (b'\xFF\xD8\xFF\xE0', b'\xFF\xD9', 'jpg', 4),  # JFIF
            (b'\xFF\xD8\xFF\xE1', b'\xFF\xD9', 'jpg', 4),  # EXIF
            (b'\xFF\xD8\xFF\xDB', b'\xFF\xD9', 'jpg', 4),  # Standard JPEG
            (b'\x89PNG\r\n\x1A\n', b'IEND\xaeB`\x82', 'png', 8),  # PNG
            (b'GIF87a', b'\x00\x3B', 'gif', 6),  # GIF87a
            (b'GIF89a', b'\x00\x3B', 'gif', 6),  # GIF89a
            (b'BM', None, 'bmp', 2),  # BMP
        ]

        found_positions = set()

        # Direct image signature search
        for start_sig, end_sig, ext, min_header_size in image_signatures:
            pos = 0
            while True:
                pos = data.find(start_sig, pos)
                if pos == -1:
                    break
                
                if pos in found_positions:
                    pos += len(start_sig)
                    continue

                try:
                    image_data = None
                    
                    if ext == 'jpg':
                        # JPEG extraction
                        end_pos = pos + len(start_sig)
                        while end_pos < len(data) - 1:
                            if data[end_pos] == 0xFF and data[end_pos + 1] == 0xD9:
                                end_pos += 2
                                image_data = data[pos:end_pos]
                                break
                            end_pos += 1
                            if end_pos - pos > 10 * 1024 * 1024:  # 10MB max
                                break
                                
                    elif ext == 'png':
                        # PNG extraction
                        end_pos = data.find(end_sig, pos)
                        if end_pos != -1:
                            end_pos += len(end_sig)
                            image_data = data[pos:end_pos]
                            
                    elif ext == 'gif':
                        # GIF extraction
                        end_pos = data.find(end_sig, pos)
                        if end_pos != -1:
                            end_pos += len(end_sig)
                            image_data = data[pos:end_pos]
                            
                    elif ext == 'bmp':
                        # BMP extraction
                        if pos + 18 <= len(data):
                            try:
                                file_size = struct.unpack('<I', data[pos + 2:pos + 6])[0]
                                if 54 <= file_size <= 50 * 1024 * 1024 and pos + file_size <= len(data):
                                    image_data = data[pos:pos + file_size]
                            except struct.error:
                                pass

                    # Validate and save image
                    if image_data and len(image_data) >= min_header_size:
                        if validate_and_save_image_improved(image_data, output_dir, ext, count):
                            found_positions.add(pos)
                            extracted_files.append(f"{output_dir}/extracted_image_{count + 1:03d}.{ext}")
                            count += 1

                except Exception as e:
                    print(f"Error processing {ext} at position {pos}: {e}")

                pos += 1

        return count, extracted_files, None if count > 0 else "No images found using binary extraction"

    except Exception as e:
        return 0, [], f"Error reading .xls file: {str(e)}"


def validate_and_save_image_improved(image_data, output_dir, ext, count):
    """
    Improved image validation and saving
    """
    try:
        img_buffer = io.BytesIO(image_data)
        
        with PILImage.open(img_buffer) as img:
            img.verify()
            
        img_buffer.seek(0)
        with PILImage.open(img_buffer) as img:
            width, height = img.size
            
            # Validate dimensions
            if width < 1 or height < 1 or width > 20000 or height > 20000:
                return False
            
            if width * height < 16:  # Less than 4x4 pixels
                return False

            # Convert problematic modes
            if img.mode in ('P', 'LA'):
                img = img.convert('RGBA')
            elif img.mode == '1':
                img = img.convert('L')

            # Handle transparency for JPEG
            if ext == 'jpg' and img.mode in ('RGBA', 'LA'):
                background = PILImage.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'RGBA':
                    background.paste(img, mask=img.split()[-1])
                else:
                    background.paste(img.convert('RGBA'), mask=img.split()[-1])
                img = background

            # Save image
            output_path = f"{output_dir}/extracted_image_{count + 1:03d}.{ext}"
            
            if ext == 'jpg':
                img.save(output_path, 'JPEG', quality=95, optimize=True)
            elif ext == 'png':
                img.save(output_path, 'PNG', optimize=True)
            elif ext == 'gif':
                if img.mode != 'P':
                    img = img.convert('P', palette=PILImage.ADAPTIVE)
                img.save(output_path, 'GIF', optimize=True)
            else:
                img.save(output_path)

            return True

    except Exception as e:
        print(f"Image validation failed: {e}")
        return False


def extract_images_from_excel_with_worksheet_improved(excel_file_path, output_dir, target_worksheet=None):
    """
    Main improved extraction function with proper worksheet support
    """
    try:
        file_ext = Path(excel_file_path).suffix.lower()
        
        if file_ext == '.xlsx':
            return extract_images_from_xlsx_worksheet_specific(excel_file_path, output_dir, target_worksheet)
        elif file_ext == '.xls':
            return extract_images_from_xls_worksheet_specific(excel_file_path, output_dir, target_worksheet)
        else:
            return 0, [], f"Unsupported file format: {file_ext}", []

    except Exception as e:
        return 0, [], str(e), []


def get_excel_worksheets(excel_file_path):
    """Get list of worksheet names from Excel file"""
    try:
        file_ext = Path(excel_file_path).suffix.lower()
        
        if file_ext == '.xlsx':
            try:
                from openpyxl import load_workbook
                wb = load_workbook(excel_file_path, read_only=True)
                worksheets = wb.sheetnames
                wb.close()
                return worksheets
            except ImportError:
                # Fallback method using zipfile
                with zipfile.ZipFile(excel_file_path, 'r') as z:
                    try:
                        workbook_xml = z.read('xl/workbook.xml').decode('utf-8')
                        root = ET.fromstring(workbook_xml)
                        sheets = root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet')
                        return [sheet.get('name') for sheet in sheets]
                    except:
                        return ["Sheet1"]
                        
        elif file_ext == '.xls':
            try:
                import xlrd
                workbook = xlrd.open_workbook(excel_file_path)
                return workbook.sheet_names()
            except ImportError:
                return ["All worksheets"]
        
        return []
        
    except Exception:
        return []

def extract_images_from_xlsx_improved(xlsx_path, output_dir, target_sheets=None):
    """
    Improved extraction of images from .xlsx files with proper sheet mapping
    """
    try:
        extracted_files = []
        image_count = 0
        
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            # Get all files in the archive
            file_list = zip_ref.namelist()
            
            # Find media files
            media_files = [f for f in file_list if f.startswith('xl/media/')]
            if not media_files:
                return 0, [], "No images found in the Excel file", []
            
            # Get worksheet information
            sheet_mapping = {}  # Maps sheet IDs to names
            sheet_rids = {}     # Maps sheet names to rIds
            
            if 'xl/workbook.xml' in file_list:
                with zip_ref.open('xl/workbook.xml') as f:
                    content = f.read().decode('utf-8')
                    import re
                    # Extract sheet information
                    sheet_matches = re.findall(r'<sheet name="([^"]+)" sheetId="(\d+)" r:id="(rId\d+)"', content)
                    for name, sheet_id, r_id in sheet_matches:
                        sheet_mapping[sheet_id] = name
                        sheet_rids[name] = r_id
            
            available_sheets = list(sheet_mapping.values())
            
            # If target_sheets is None, extract from all sheets
            if target_sheets is None:
                target_sheets = available_sheets
            elif isinstance(target_sheets, str):
                target_sheets = [target_sheets] if target_sheets != "All worksheets" else available_sheets
            
            # Map images to sheets by analyzing drawing relationships
            sheet_images = {}  # Maps sheet names to their image files
            
            # Check each worksheet for drawing relationships
            for sheet_id, sheet_name in sheet_mapping.items():
                if sheet_name not in target_sheets:
                    continue
                    
                sheet_images[sheet_name] = []
                
                # Look for drawing relationships for this sheet
                drawing_rels_file = f'xl/worksheets/_rels/sheet{sheet_id}.xml.rels'
                if drawing_rels_file not in file_list:
                    # Try alternative naming patterns
                    possible_names = [
                        f'xl/worksheets/_rels/sheet{sheet_id}.xml.rels',
                        f'xl/worksheets/_rels/sheet{int(sheet_id)}.xml.rels'
                    ]
                    drawing_rels_file = None
                    for possible in possible_names:
                        if possible in file_list:
                            drawing_rels_file = possible
                            break
                
                if drawing_rels_file:
                    try:
                        with zip_ref.open(drawing_rels_file) as f:
                            rels_content = f.read().decode('utf-8')
                        
                        # Find drawing references
                        drawing_matches = re.findall(r'Target="([^"]*drawing\d+\.xml)"', rels_content)
                        
                        for drawing_file in drawing_matches:
                            drawing_path = f'xl/{drawing_file}'
                            if drawing_path in file_list:
                                # Now get the drawing relationships
                                drawing_rels_path = drawing_path.replace('.xml', '.xml.rels').replace('xl/', 'xl/').replace('drawings/', 'drawings/_rels/')
                                
                                if drawing_rels_path in file_list:
                                    with zip_ref.open(drawing_rels_path) as f:
                                        drawing_rels_content = f.read().decode('utf-8')
                                    
                                    # Extract image references
                                    image_matches = re.findall(r'Target="([^"]*media/[^"]+)"', drawing_rels_content)
                                    for img_ref in image_matches:
                                        img_path = f'xl/drawings/{img_ref}' if not img_ref.startswith('../') else f'xl/{img_ref[3:]}'
                                        if img_path in file_list:
                                            sheet_images[sheet_name].append(img_path)
                    except Exception as e:
                        print(f"Error processing relationships for sheet {sheet_name}: {e}")
            
            # If no sheet-specific images found, fall back to extracting all images
            if not any(sheet_images.values()):
                print("No sheet-specific images found, extracting all images")
                for sheet_name in target_sheets:
                    sheet_images[sheet_name] = media_files.copy()
            
            # Extract images for target sheets
            extracted_media = set()  # Avoid duplicates
            
            for sheet_name in target_sheets:
                if sheet_name in sheet_images:
                    for media_file in sheet_images[sheet_name]:
                        if media_file not in extracted_media:
                            try:
                                # Extract the image
                                image_data = zip_ref.read(media_file)
                                
                                # Get original filename and extension
                                original_filename = os.path.basename(media_file)
                                name, ext = os.path.splitext(original_filename)
                                
                                # Create output filename with sheet context
                                if len(target_sheets) > 1:
                                    output_filename = f"{sheet_name}_{original_filename}"
                                else:
                                    output_filename = original_filename
                                
                                output_path = os.path.join(output_dir, output_filename)
                                
                                # Save the image
                                with open(output_path, 'wb') as img_file:
                                    img_file.write(image_data)
                                
                                extracted_files.append(output_path)
                                extracted_media.add(media_file)
                                image_count += 1
                                
                            except Exception as e:
                                print(f"Error extracting {media_file}: {e}")
        
        return image_count, extracted_files, None, available_sheets
        
    except Exception as e:
        return 0, [], f"Error processing Excel file: {str(e)}", []



def get_excel_worksheets_simple(excel_file_path):
    """Get list of worksheet names from Excel file - simplified version"""
    try:
        file_ext = Path(excel_file_path).suffix.lower()
        
        if file_ext == '.xlsx':
            try:
                workbook = openpyxl.load_workbook(excel_file_path, read_only=True)
                worksheets = workbook.sheetnames
                workbook.close()
                return worksheets
            except Exception:
                return ["Sheet1"]  # Fallback
                        
        elif file_ext == '.xls':
            try:
                import xlrd
                workbook = xlrd.open_workbook(excel_file_path)
                return workbook.sheet_names()
            except ImportError:
                return ["Sheet1"]  # Fallback instead of "All worksheets"
            except Exception:
                return ["Sheet1"]
        
        return []
        
    except Exception:
        return []

# Add the missing extract_ole_images_from_xls_comprehensive function that was referenced
def extract_ole_images_from_xls_comprehensive(xls_file_path, output_dir):
    """
    Comprehensive extraction of embedded images from .xls files using binary parsing
    """
    extracted_files = []
    count = 0
    
    try:
        # Read the entire file as binary
        with open(xls_file_path, 'rb') as f:
            data = f.read()

        # Enhanced image signatures
        image_signatures = [
            (b'\xFF\xD8\xFF\xE0', b'\xFF\xD9', 'jpg', 4),  # JFIF
            (b'\xFF\xD8\xFF\xE1', b'\xFF\xD9', 'jpg', 4),  # EXIF
            (b'\xFF\xD8\xFF\xDB', b'\xFF\xD9', 'jpg', 4),  # Standard JPEG
            (b'\x89PNG\r\n\x1A\n', b'IEND\xaeB`\x82', 'png', 8),  # PNG
            (b'GIF87a', b'\x00\x3B', 'gif', 6),  # GIF87a
            (b'GIF89a', b'\x00\x3B', 'gif', 6),  # GIF89a
            (b'BM', None, 'bmp', 2),  # BMP
        ]

        found_positions = set()

        # Direct image signature search
        for start_sig, end_sig, ext, min_header_size in image_signatures:
            pos = 0
            while True:
                pos = data.find(start_sig, pos)
                if pos == -1:
                    break
                
                if pos in found_positions:
                    pos += len(start_sig)
                    continue

                try:
                    image_data = None
                    
                    if ext == 'jpg':
                        # JPEG extraction
                        end_pos = pos + len(start_sig)
                        while end_pos < len(data) - 1:
                            if data[end_pos] == 0xFF and data[end_pos + 1] == 0xD9:
                                end_pos += 2
                                image_data = data[pos:end_pos]
                                break
                            end_pos += 1
                            if end_pos - pos > 10 * 1024 * 1024:  # 10MB max
                                break
                                
                    elif ext == 'png':
                        # PNG extraction
                        end_pos = data.find(end_sig, pos)
                        if end_pos != -1:
                            end_pos += len(end_sig)
                            image_data = data[pos:end_pos]
                            
                    elif ext == 'gif':
                        # GIF extraction
                        end_pos = data.find(end_sig, pos)
                        if end_pos != -1:
                            end_pos += len(end_sig)
                            image_data = data[pos:end_pos]
                            
                    elif ext == 'bmp':
                        # BMP extraction
                        if pos + 18 <= len(data):
                            try:
                                import struct
                                file_size = struct.unpack('<I', data[pos + 2:pos + 6])[0]
                                if 54 <= file_size <= 50 * 1024 * 1024 and pos + file_size <= len(data):
                                    image_data = data[pos:pos + file_size]
                            except:
                                pass

                    # Validate and save image
                    if image_data and len(image_data) >= min_header_size:
                        if validate_and_save_image_improved(image_data, output_dir, ext, count):
                            found_positions.add(pos)
                            extracted_files.append(f"{output_dir}/extracted_image_{count + 1:03d}.{ext}")
                            count += 1

                except Exception as e:
                    print(f"Error processing {ext} at position {pos}: {e}")

                pos += 1

        return count, extracted_files, None if count > 0 else "No images found using binary extraction"

    except Exception as e:
        return 0, [], f"Error reading .xls file: {str(e)}"

def validate_and_save_image_improved(image_data, output_dir, ext, count):
    """
    Improved image validation and saving
    """
    try:
        img_buffer = io.BytesIO(image_data)
        
        with PILImage.open(img_buffer) as img:
            img.verify()
            
        img_buffer.seek(0)
        with PILImage.open(img_buffer) as img:
            width, height = img.size
            
            # Validate dimensions
            if width < 1 or height < 1 or width > 20000 or height > 20000:
                return False
            
            if width * height < 16:  # Less than 4x4 pixels
                return False

            # Convert problematic modes
            if img.mode in ('P', 'LA'):
                img = img.convert('RGBA')
            elif img.mode == '1':
                img = img.convert('L')

            # Handle transparency for JPEG
            if ext == 'jpg' and img.mode in ('RGBA', 'LA'):
                background = PILImage.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'RGBA':
                    background.paste(img, mask=img.split()[-1])
                else:
                    background.paste(img.convert('RGBA'), mask=img.split()[-1])
                img = background

            # Save image
            output_path = f"{output_dir}/extracted_image_{count + 1:03d}.{ext}"
            
            if ext == 'jpg':
                img.save(output_path, 'JPEG', quality=95, optimize=True)
            elif ext == 'png':
                img.save(output_path, 'PNG', optimize=True)
            elif ext == 'gif':
                if img.mode != 'P':
                    img = img.convert('P', palette=PILImage.ADAPTIVE)
                img.save(output_path, 'GIF', optimize=True)
            else:
                img.save(output_path)

            return True

    except Exception as e:
        print(f"Image validation failed: {e}")
        return False

def extract_images_from_xlsx_by_sheet(xlsx_path, output_dir, target_sheet=None):
    """
    Extract images from .xlsx files with reliable sheet-specific extraction
    Uses openpyxl to directly access worksheet images
    """
    try:
        extracted_files = []
        image_count = 0
        
        # Load workbook with openpyxl
        workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
        available_sheets = workbook.sheetnames
        
        # Determine which sheets to process
        if target_sheet is None or target_sheet == "All worksheets":
            sheets_to_process = available_sheets
        else:
            sheets_to_process = [target_sheet] if target_sheet in available_sheets else []
        
        if not sheets_to_process:
            workbook.close()
            return 0, [], "No valid sheets found to extract from", available_sheets
        
        # Process each sheet
        for sheet_name in sheets_to_process:
            try:
                worksheet = workbook[sheet_name]
                
                # Check if worksheet has any images
                if hasattr(worksheet, '_images') and worksheet._images:
                    sheet_image_count = 0
                    for image in worksheet._images:
                        try:
                            # Get image data
                            if hasattr(image, '_data'):
                                image_data = image._data()
                            else:
                                # Try alternative method
                                continue
                            
                            # Determine file extension from image format
                            image_format = 'png'  # default
                            if hasattr(image, 'format') and image.format:
                                image_format = image.format.lower()
                            elif len(image_data) > 4:
                                # Try to detect format from data
                                if image_data[:4] == b'\x89PNG':
                                    image_format = 'png'
                                elif image_data[:3] == b'\xFF\xD8\xFF':
                                    image_format = 'jpg'
                                elif image_data[:6] == b'GIF87a' or image_data[:6] == b'GIF89a':
                                    image_format = 'gif'
                            
                            # Create filename
                            if len(sheets_to_process) > 1:
                                filename = f"{sheet_name}_image_{sheet_image_count + 1:03d}.{image_format}"
                            else:
                                filename = f"image_{sheet_image_count + 1:03d}.{image_format}"
                            
                            # Save image
                            output_path = os.path.join(output_dir, filename)
                            with open(output_path, 'wb') as f:
                                f.write(image_data)
                            
                            extracted_files.append(output_path)
                            sheet_image_count += 1
                            image_count += 1
                            
                        except Exception as e:
                            print(f"Error extracting image from {sheet_name}: {e}")
                            continue
                    
                    print(f"Extracted {sheet_image_count} images from sheet: {sheet_name}")
                
                # Also check for drawing objects (alternative method)
                if hasattr(worksheet, 'drawing') and worksheet.drawing:
                    for drawing in worksheet.drawing:
                        if hasattr(drawing, '_images'):
                            for img in drawing._images:
                                try:
                                    # Similar extraction logic for drawing images
                                    if hasattr(img, '_data'):
                                        image_data = img._data()
                                        image_format = 'png'
                                        
                                        if len(sheets_to_process) > 1:
                                            filename = f"{sheet_name}_drawing_{image_count + 1:03d}.{image_format}"
                                        else:
                                            filename = f"drawing_{image_count + 1:03d}.{image_format}"
                                        
                                        output_path = os.path.join(output_dir, filename)
                                        with open(output_path, 'wb') as f:
                                            f.write(image_data)
                                        
                                        extracted_files.append(output_path)
                                        image_count += 1
                                        
                                except Exception:
                                    continue
                            
            except Exception as e:
                print(f"Error processing sheet {sheet_name}: {e}")
                continue
        
        workbook.close()
        
        # If openpyxl method didn't find images, fall back to zip extraction
        if image_count == 0:
            return extract_images_from_xlsx_zip_method(xlsx_path, output_dir, target_sheet)
        
        return image_count, extracted_files, None, available_sheets
        
    except Exception as e:
        return extract_images_from_xlsx_zip_method(xlsx_path, output_dir, target_sheet)

def extract_images_from_xlsx_zip_method(xlsx_path, output_dir, target_sheet=None):
    """
    Fallback method: Extract all images from xlsx using zip method
    """
    try:
        extracted_files = []
        image_count = 0
        
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            # Get available sheets
            available_sheets = []
            try:
                with zip_ref.open('xl/workbook.xml') as f:
                    content = f.read().decode('utf-8')
                sheet_matches = re.findall(r'<sheet name="([^"]+)"', content)
                available_sheets = [match for match in sheet_matches]
            except:
                available_sheets = ["Sheet1"]
            
            # Find all media files
            media_files = [f for f in zip_ref.namelist() if f.startswith('xl/media/')]
            
            if not media_files:
                return 0, [], "No images found in Excel file", available_sheets
            
            # Extract all images (since precise sheet mapping is complex)
            for media_file in media_files:
                try:
                    image_data = zip_ref.read(media_file)
                    original_filename = os.path.basename(media_file)
                    
                    # Create output filename
                    if target_sheet and target_sheet != "All worksheets":
                        output_filename = f"{target_sheet}_{original_filename}"
                    else:
                        output_filename = original_filename
                    
                    output_path = os.path.join(output_dir, output_filename)
                    
                    with open(output_path, 'wb') as f:
                        f.write(image_data)
                    
                    extracted_files.append(output_path)
                    image_count += 1
                    
                except Exception as e:
                    print(f"Error extracting {media_file}: {e}")
                    continue
        
        warning_msg = None
        if target_sheet and target_sheet != "All worksheets":
            warning_msg = f"Note: Extracted all images from file. Sheet-specific extraction not fully supported with this method."
        
        return image_count, extracted_files, warning_msg, available_sheets
        
    except Exception as e:
        return 0, [], f"Error processing Excel file: {str(e)}", []

def excel_extractor_tool():
    st.header("Excel Image Extractor")
    st.markdown("Upload Excel files (.xlsx) to extract embedded images from specific sheets.")
    
    # File upload for Excel files
    uploaded_files = st.file_uploader(
        "Choose Excel file(s)",
        type=['xlsx'],
        accept_multiple_files=True,
        help="Upload one or more .xlsx files to extract embedded images.",
        key="excel_upload_fixed"
    )
    
    if uploaded_files:
        # For each file, let user select sheets
        sheet_selections = {}
        
        for uploaded_file in uploaded_files:
            # Temporarily save file to read sheet names
            temp_dir = tempfile.mkdtemp()
            temp_path = os.path.join(temp_dir, uploaded_file.name)
            
            with open(temp_path, 'wb') as f:
                f.write(uploaded_file.getbuffer())
            
            try:
                # Get available sheets using the simple function
                available_sheets = get_excel_worksheets_simple(temp_path)
                
                if available_sheets:
                    st.subheader(f"Select sheets from: {uploaded_file.name}")
                    
                    # Add "All sheets" option
                    all_options = ["All worksheets"] + available_sheets
                    
                    selected = st.multiselect(
                        f"Choose sheets to extract images from:",
                        all_options,
                        default=["All worksheets"],
                        help="Select specific sheets or 'All worksheets' to extract from all sheets",
                        key=f"sheets_fixed_{uploaded_file.name}"
                    )
                    
                    if "All worksheets" in selected:
                        sheet_selections[uploaded_file.name] = available_sheets
                    else:
                        sheet_selections[uploaded_file.name] = [s for s in selected if s != "All worksheets"]
                else:
                    st.warning(f"Could not read sheets from {uploaded_file.name}")
                    sheet_selections[uploaded_file.name] = []
                    
            except Exception as e:
                st.error(f"Error reading {uploaded_file.name}: {str(e)}")
                sheet_selections[uploaded_file.name] = []
            
            # Clean up
            try:
                os.remove(temp_path)
                os.rmdir(temp_dir)
            except:
                pass
        
        if st.button("Extract Images", type="primary", key="extract_btn_fixed"):
            temp_dir = tempfile.mkdtemp()
            all_extracted_files = []
            total_images = 0
            extraction_results = []
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            for i, uploaded_file in enumerate(uploaded_files):
                file_name = uploaded_file.name
                selected_sheets = sheet_selections.get(file_name, [])
                
                if not selected_sheets:
                    st.warning(f"No sheets selected for {file_name}, skipping...")
                    continue
                
                status_text.text(f"Processing {file_name}...")
                progress_bar.progress((i + 1) / len(uploaded_files))
                
                # Save uploaded file temporarily
                temp_xlsx_path = os.path.join(temp_dir, file_name)
                with open(temp_xlsx_path, 'wb') as f:
                    f.write(uploaded_file.getbuffer())
                
                # Create output directory for this file
                output_dir = os.path.join(temp_dir, f"{Path(file_name).stem}_images")
                os.makedirs(output_dir, exist_ok=True)
                
                # Extract images for each selected sheet individually
                file_total = 0
                for sheet_name in selected_sheets:
                    sheet_output_dir = os.path.join(output_dir, sheet_name.replace('/', '_').replace('\\', '_'))
                    os.makedirs(sheet_output_dir, exist_ok=True)
                    
                    count, extracted_files, error, _ = extract_images_from_xlsx_by_sheet(
                        temp_xlsx_path, 
                        sheet_output_dir, 
                        sheet_name
                    )
                    
                    if error:
                        st.warning(f"Issue with {file_name} - {sheet_name}: {error}")
                    
                    if count > 0:
                        extraction_results.append(f"✅ {file_name} - {sheet_name}: {count} images")
                        all_extracted_files.extend(extracted_files)
                        file_total += count
                    else:
                        extraction_results.append(f"⚪ {file_name} - {sheet_name}: No images found")
                
                total_images += file_total
                if file_total > 0:
                    st.success(f"Extracted {file_total} images from {file_name}")
                else:
                    st.info(f"No images found in {file_name}")
            
            status_text.empty()
            progress_bar.empty()
            
            # Show detailed extraction results
            with st.expander("Detailed Extraction Results"):
                for result in extraction_results:
                    st.write(result)
            
            if total_images > 0:
                st.success(f"Total: {total_images} images extracted from {len(uploaded_files)} file(s)")
                
                # Show extracted files summary
                with st.expander("Extracted Files Summary"):
                    for file_path in all_extracted_files[:20]:  # Show first 20
                        st.write(f"📄 {os.path.relpath(file_path, temp_dir)}")
                    if len(all_extracted_files) > 20:
                        st.info(f"Showing first 20 files. Total extracted: {len(all_extracted_files)}")
                
                # Create zip file with all extracted images
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    for file_path in all_extracted_files:
                        # Get relative path for zip structure
                        rel_path = os.path.relpath(file_path, temp_dir)
                        zip_file.write(file_path, rel_path)
                
                zip_buffer.seek(0)
                
                # Download button for all extracted images
                st.download_button(
                    label="Download All Extracted Images (ZIP)",
                    data=zip_buffer.getvalue(),
                    file_name="extracted_images_by_sheet.zip",
                    mime="application/zip"
                )
            
            else:
                st.info("No images were found in the selected sheets of the uploaded Excel file(s).")
    
    else:
        st.info("Please upload Excel file(s) to get started.")
        
        with st.expander("Instructions"):
            st.markdown("""
 
            """)

def file_renamer_tool():
    st.header("Brand File Renamer")
    st.markdown("Upload a zip file containing brand folders to automatically rename files with brand prefixes.")
    
    # File upload
    uploaded_file = st.file_uploader(
        "Choose a ZIP file",
        type=['zip'],
        # help="Upload a zip file containing your brand folders with assets.",
        key="renamer_zip_upload"
    )
    
    if uploaded_file is not None:
        try:
            # Extract zip file
            with st.spinner("Extracting zip file..."):
                temp_dir = extract_zip_to_temp(uploaded_file)
            
            # Find the actual input folder (in case zip has a root folder)
            input_folder = Path(temp_dir)
            items = os.listdir(temp_dir)
            if len(items) == 1 and os.path.isdir(os.path.join(temp_dir, items[0])):
                input_folder = Path(temp_dir) / items[0]
            
            # Create output folder
            output_folder = input_folder.parent / (input_folder.name + "_renamed")
            output_folder.mkdir(exist_ok=True)
            
            # Process files
            total_files = 0
            processed_files = []
            file_type_counts = {"images": 0, "text": 0, "videos": 0}
            
            for subfolder in input_folder.iterdir():
                if subfolder.is_dir():
                    markenname = subfolder.name
                    new_subfolder = output_folder / markenname
                    new_subfolder.mkdir(parents=True, exist_ok=True)
                    
                    for file in subfolder.iterdir():
                        if file.is_file() and file.suffix.lower() in ALL_ALLOWED_EXTENSIONS:
                            new_name = f"{markenname}_{file.name}"
                            target_path = new_subfolder / new_name
                            shutil.copy2(file, target_path)
                            processed_files.append((file.name, new_name, markenname))
                            total_files += 1
                            
                            # Count file types
                            if file.suffix.lower() in ALLOWED_IMAGE_EXTENSIONS:
                                file_type_counts["images"] += 1
                            elif file.suffix.lower() in ALLOWED_TEXT_EXTENSIONS:
                                file_type_counts["text"] += 1
                            elif file.suffix.lower() in ALLOWED_VIDEO_EXTENSIONS:
                                file_type_counts["videos"] += 1
            
            if total_files > 0:
                st.success(f"Successfully renamed and copied {total_files} files")
                
                # Show file type breakdown
                st.info(f"File types processed: {file_type_counts['images']} images, {file_type_counts['text']} text files, {file_type_counts['videos']} videos")
                
                # Show preview
                with st.expander("📋 Preview renamed files"):
                    preview_df = pd.DataFrame(
                        processed_files,  
                        columns=["Original Name", "New Name", "Brand"]
                    )
                    st.dataframe(preview_df)
                    
                    if total_files > 20:
                        st.info(f"Total processed: {total_files}")
                
                # Create zip file with renamed files
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    for root, dirs, files in os.walk(output_folder):
                        for file in files:
                            file_path = Path(root) / file
                            rel_path = os.path.relpath(file_path, output_folder)
                            zip_file.write(file_path, rel_path)
                
                zip_buffer.seek(0)
                
                # Download button
                st.download_button(
                    label="📥 Download Renamed Files (ZIP)",
                    data=zip_buffer.getvalue(),
                    file_name="brand_files_renamed.zip",
                    mime="application/zip"
                )
            else:
                st.warning("No files were found to rename in the uploaded zip file.")
        
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
    
    else:
        st.info("👆 Please upload a zip file to get started.")
        
        # Instructions
        with st.expander("📖 Instructions"):
            st.markdown("""
            **How to use this tool:**
            
            1. Create folders named after your brands
            2. Place brand assets inside each brand folder
            3. Zip the entire structure
            4. Upload the zip file
            
            **The tool will:**
            - Create a new folder structure matching the original
            - Rename all files with the brand prefix (e.g., "Brand1_filename.jpg")
            - Provide a zip file with all renamed files
            
            **Supported file types:**
            - **Images:** JPG, JPEG, PNG, BMP, GIF, TIFF, WEBP
            - **Text files:** TXT, MD, CSV
            - **Videos:** MP4, AVI, MOV, WMV, FLV, MKV, WEBM, M4V, 3GP, OGV
            """)



def white_to_transparent_tool():
    st.header("White Background to Transparent")
    st.markdown("Upload images to convert white backgrounds to transparent.")
    
    # Options
    col1, col2 = st.columns(2)
    with col1:
        crop_image = st.checkbox("Crop white background", value=False)
    with col2:
        overwrite = st.checkbox("Overwrite original files", value=False)
    
    # File upload
    uploaded_files = st.file_uploader(
        "Choose image files",
        type=list(ALLOWED_IMAGE_EXTENSIONS),
        accept_multiple_files=True,
        help="Upload images to convert white backgrounds to transparent.",
        key="transparent_upload"
    )
    
    if uploaded_files:
        if st.button("Process Images", type="primary"):
            temp_dir = tempfile.mkdtemp()
            processed_files = []
            failed_files = []
            
            # Save uploaded files to temp dir
            for uploaded_file in uploaded_files:
                file_path = os.path.join(temp_dir, uploaded_file.name)
                with open(file_path, 'wb') as f:
                    f.write(uploaded_file.getbuffer())
            
            # Process images
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            for i, uploaded_file in enumerate(uploaded_files):
                status_text.text(f"Processing {uploaded_file.name}... ({i+1}/{len(uploaded_files)})")
                progress_bar.progress((i + 1) / len(uploaded_files))
                
                try:
                    img = PILImage.open(os.path.join(temp_dir, uploaded_file.name)).convert("RGBA")
                    
                    # Convert to numpy array for OpenCV processing
                    img_array = np.array(img)
                    
                    if crop_image:
                        # Crop image
                        gray = cv2.cvtColor(img_array, cv2.COLOR_BGR2GRAY)
                        th, threshed = cv2.threshold(gray, 240, 255, cv2.THRESH_BINARY_INV)
                        
                        kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (11,11))
                        morphed = cv2.morphologyEx(threshed, cv2.MORPH_CLOSE, kernel)
                        
                        cnts, _ = cv2.findContours(morphed, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                        if cnts:  # Only crop if contours are found
                            cnt = sorted(cnts, key=cv2.contourArea)[-1]
                            x,y,w,h = cv2.boundingRect(cnt)
                            img_array = img_array[y:y+h, x:x+w]
                    
                    # Make white background transparent
                    gray = cv2.cvtColor(img_array, cv2.COLOR_BGR2GRAY)
                    th, threshed = cv2.threshold(gray, 240, 255, cv2.THRESH_BINARY_INV)
                    
                    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (11,11))
                    morphed = cv2.morphologyEx(threshed, cv2.MORPH_CLOSE, kernel)
                    
                    roi, _ = cv2.findContours(morphed, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    
                    mask = np.zeros(img_array.shape, img_array.dtype)
                    if roi:  # Only proceed if contours are found
                        cv2.fillPoly(mask, roi, (255,)*img_array.shape[2])
                        masked_image = cv2.bitwise_and(img_array, mask)
                        
                        # Convert back to PIL Image
                        result_img = PILImage.fromarray(masked_image, mode="RGBA")
                        
                        # Save the processed image
                        new_filename = os.path.splitext(uploaded_file.name)[0] + ".png"
                        processed_path = os.path.join(temp_dir, 
                                                       new_filename)
                        result_img.save(processed_path)
                        processed_files.append(processed_path)
                    else:
                        failed_files.append(uploaded_file.name)
                except Exception as e:
                    failed_files.append(uploaded_file.name)
                    st.warning(f"Failed to process {uploaded_file.name}: {str(e)}")
            
            status_text.empty()
            progress_bar.empty()
            
            if processed_files:
                st.success(f"✅ Successfully processed {len(processed_files)} images")
                
                # Create zip file with processed images
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    for file_path in processed_files:
                        zip_file.write(file_path, os.path.basename(file_path))
                
                zip_buffer.seek(0)
                
                # Download button
                st.download_button(
                    label="📥 Download Processed Images (ZIP)",
                    data=zip_buffer.getvalue(),
                    file_name="transparent_images.zip",
                    mime="application/zip"
                )
                
                # Show preview
                with st.expander("🖼️ Preview Processed Images"):
                    cols = st.columns(3)
                    for i, file_path in enumerate(processed_files[:6]):  # Show max 6 images
                        with cols[i % 3]:
                            img = PILImage.open(file_path)
                            st.image(img, caption=os.path.basename(file_path), use_column_width=True)
                    
                    if len(processed_files) > 6:
                        st.info(f"Showing first 6 images. Total processed: {len(processed_files)}")
            
            if failed_files:
                st.warning(f"⚠️ Failed to process {len(failed_files)} files: {', '.join(failed_files)}")
    
    else:
        st.info("👆 Please upload image files to get started.")
        
        # Instructions
        with st.expander("📖 Instructions"):
            st.markdown("""
            **How to use this tool:**
            
            1. Upload one or more image files
            2. Choose processing options:
               - Crop white background: Removes excess white space around the image
               - Overwrite original files: Not applicable in this web version (always creates new files)
            3. Click "Process Images"
            4. Download the processed images with transparent backgrounds
            
            **Note:**
            - Output images will always be in PNG format to support transparency
            - Only white backgrounds will be made transparent (threshold of 240/255)
            - Complex images may require manual adjustment for best results
            """)

def find_smallest_dimensions_tool():
    st.header("Find Smallest Image Dimensions")
    st.markdown("Analyze a folder to find images with the smallest width and height.")
    
    # File upload
    uploaded_file = st.file_uploader(
        "Choose a ZIP file with images",
        type=['zip'],
        help="Upload a zip file containing images to analyze.",
        key="dimensions_zip_upload"
    )
    
    if uploaded_file:
        try:
            # Extract zip file
            with st.spinner("Extracting zip file..."):
                temp_dir = extract_zip_to_temp(uploaded_file)
            
            # Find the actual input folder (in case zip has a root folder)
            input_folder = temp_dir
            items = os.listdir(temp_dir)
            if len(items) == 1 and os.path.isdir(os.path.join(temp_dir, items[0])):
                input_folder = os.path.join(temp_dir, items[0])
            
            # Analyze images
            with st.spinner("Analyzing images..."):
                min_width = None
                min_height = None
                file_min_width = ""
                file_min_height = ""
                height_at_min_width = None
                width_at_min_height = None
                total_images = 0
                
                for root, dirs, files in os.walk(input_folder):
                    for file in files:
                        ext = os.path.splitext(file)[1].lower()
                        if ext in ALLOWED_IMAGE_EXTENSIONS:
                            path = os.path.join(root, file)
                            try:
                                with PILImage.open(path) as img:
                                    width, height = img.size
                                    total_images += 1
                                    
                                    if (min_width is None) or (width < min_width):
                                        min_width = width
                                        height_at_min_width = height
                                        file_min_width = path
                                    
                                    if (min_height is None) or (height < min_height):
                                        min_height = height
                                        width_at_min_height = width
                                        file_min_height = path
                            except Exception as e:
                                st.warning(f"Error processing {file}: {str(e)}")
                
                if min_width is not None and min_height is not None:
                    st.success(f"Analyzed {total_images} images")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        st.subheader("Smallest Width")
                        st.write(f"Dimensions: {min_width} × {height_at_min_width} px")
                        st.image(file_min_width, caption=os.path.basename(file_min_width), use_column_width =True)
                    
                    with col2:
                        st.subheader("Smallest Height")
                        st.write(f"Dimensions: {width_at_min_height} × {min_height} px")
                        st.image(file_min_height, caption=os.path.basename(file_min_height), use_column_width =True)
                else:
                    st.warning("No valid image files found in the uploaded zip.")
        
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
    
    else:
        st.info("👆 Please upload a zip file with images to analyze.")
        
        # Instructions
        with st.expander("📖 Instructions"):
            st.markdown("""
            **How to use this tool:**
            
            1. Upload a zip file containing images
            2. The tool will analyze all images in the folder and subfolders
            3. It will find and display:
               - The image with the smallest width
               - The image with the smallest height
            
            **Supported image formats:**
            - JPG, JPEG, PNG, BMP, GIF, TIFF, WEBP
            """)

def resize_with_transparent_canvas_tool():
    st.header("Resize with Transparent Canvas")
    st.markdown("Resize images to 400px on the largest side while maintaining aspect ratio on transparent background.")
    
    # File upload with unique key
    uploaded_file = st.file_uploader(
        "Choose a ZIP file with images",
        type=['zip'],
        help="Upload a zip file containing images to resize.",
        key="resize_transparent_upload"
    )
    
    if uploaded_file:
        # Show the fixed resize setting
        st.info("Images will be resized to 400px on the largest side (width or height), maintaining aspect ratio.")
        
        # Process button with unique key
        if st.button("Process Images", type="primary", key="process_resize_transparent_images"):
            try:
                # Extract zip file
                with st.spinner("Extracting zip file..."):
                    temp_dir = extract_zip_to_temp(uploaded_file)
                    input_folder = temp_dir
                    items = os.listdir(temp_dir)
                    if len(items) == 1 and os.path.isdir(os.path.join(temp_dir, items[0])):
                        input_folder = os.path.join(temp_dir, items[0])
                
                # Create output folder
                output_folder = os.path.join(temp_dir, "resized_transparent_output")
                os.makedirs(output_folder, exist_ok=True)
                
                # Process images
                processed_count = 0
                error_count = 0
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                all_files = []
                for root, dirs, files in os.walk(input_folder):
                    for file in files:
                        ext = os.path.splitext(file)[1].lower()
                        if ext in ALLOWED_IMAGE_EXTENSIONS:
                            all_files.append(os.path.join(root, file))
                
                total_files = len(all_files)
                
                for i, file_path in enumerate(all_files):
                    status_text.text(f"Processing {i+1}/{total_files}: {os.path.basename(file_path)}...")
                    progress_bar.progress((i + 1) / total_files)
                    
                    try:
                        relative_path = os.path.relpath(file_path, input_folder)
                        new_path = os.path.join(output_folder, os.path.splitext(relative_path)[0] + ".png")
                        os.makedirs(os.path.dirname(new_path), exist_ok=True)
                        
                        with PILImage.open(file_path) as img:
                            img = img.convert("RGBA")
                            orig_w, orig_h = img.size
                            
                            # Calculate scaling factor based on largest side
                            max_dimension = 400
                            if orig_w >= orig_h:
                                # Width is larger or equal
                                scale_factor = max_dimension / orig_w
                                new_w = max_dimension
                                new_h = int(orig_h * scale_factor)
                            else:
                                # Height is larger
                                scale_factor = max_dimension / orig_h
                                new_h = max_dimension
                                new_w = int(orig_w * scale_factor)
                            
                            # Resize image while maintaining aspect ratio
                            img_resized = img.resize((new_w, new_h), PILImage.LANCZOS)
                            
                            # Create transparent canvas with the exact resized dimensions
                            canvas = PILImage.new("RGBA", (new_w, new_h), (0, 0, 0, 0))
                            
                            # The resized image fits perfectly on the canvas (no centering needed)
                            canvas.paste(img_resized, (0, 0), mask=img_resized)
                            
                            canvas.save(new_path)
                            processed_count += 1
                    except Exception as e:
                        error_count += 1
                        st.warning(f"Error processing {os.path.basename(file_path)}: {str(e)}")
                
                status_text.empty()
                progress_bar.empty()
                
                # Results summary
                st.success(f"""
                Processing complete!
                - ✅ Processed: {processed_count} images
                - ❌ Errors: {error_count} images
                - All images resized to 400px on largest side
                """)
                
                if processed_count > 0:
                    # Create zip file with processed images
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                        for root, dirs, files in os.walk(output_folder):
                            for file in files:
                                file_path = os.path.join(root, file)
                                rel_path = os.path.relpath(file_path, output_folder)
                                zip_file.write(file_path, rel_path)
                    
                    zip_buffer.seek(0)
                    
                    # Download button with unique key
                    st.download_button(
                        label="📥 Download Resized Images (ZIP)",
                        data=zip_buffer.getvalue(),
                        file_name="resized_400px_images.zip",
                        mime="application/zip",
                        key="download_resized_transparent"
                    )
                    
                    # Show preview
                    with st.expander("🖼️ Preview Resized Images"):
                        sample_files = []
                        for root, dirs, files in os.walk(output_folder):
                            for file in files[:3]:  # Get first 3 files from each folder
                                if len(sample_files) < 6:  # Max 6 samples
                                    sample_files.append(os.path.join(root, file))
                        
                        if sample_files:
                            cols = st.columns(3)
                            for i, file_path in enumerate(sample_files):
                                with cols[i % 3]:
                                    img = PILImage.open(file_path)
                                    st.image(img, caption=f"{os.path.basename(file_path)} ({img.size[0]}x{img.size[1]})", use_column_width=True)
                        
                        if processed_count > 6:
                            st.info(f"Showing sample images. Total processed: {processed_count}")
            
            except Exception as e:
                st.error(f"An error occurred during processing: {str(e)}")
    
    else:
        st.info("👆 Please upload a zip file with images to resize.")
        
        # Instructions
        with st.expander("📖 Instructions"):
            st.markdown("""
            **How to use this tool:**
            
            1. Upload a zip file containing images
            2. All images will be resized to 400px on their largest side (width or height)
            3. Aspect ratio is maintained - smaller dimension will be scaled proportionally
            4. Click "Process Images"
            5. Download the resized images with transparent backgrounds
            
            **Examples:**
            - 1000x800 image → 400x320 image (width was larger)
            - 600x1200 image → 200x400 image (height was larger)  
            - 400x400 image → 400x400 image (already correct size)
            
            **Features:**
            - Maintains original aspect ratio
            - Outputs PNG files to preserve transparency
            - Preserves folder structure
            - Fixed 400px maximum on largest dimension
            """)

def center_on_canvas_tool():
    st.header("Center on Transparent Canvas")
    st.markdown("Center small images on a transparent 500x500 canvas.")
    
    # File upload - add unique key
    uploaded_file = st.file_uploader(
        "Choose a ZIP file with images",
        type=['zip'],
        help="Upload a zip file containing images to process.",
        key="canvas_center_upload"  # Added unique key
    )
    
    if uploaded_file:
        # Options
        canvas_size = st.number_input(
            "Canvas size (pixels):",
            min_value=100,
            max_value=5000,
            value=500,
            step=10,
            key="canvas_size_input"  # Added unique key
        )
        
        # Add unique key to the button
        if st.button("Process Images", type="primary", key="process_canvas_images"):
            try:
                # Extract zip file
                with st.spinner("Extracting zip file..."):
                    temp_dir = extract_zip_to_temp(uploaded_file)
                    input_folder = temp_dir
                    items = os.listdir(temp_dir)
                    if len(items) == 1 and os.path.isdir(os.path.join(temp_dir, items[0])):
                        input_folder = os.path.join(temp_dir, items[0])
                
                # Create output folder
                output_folder = os.path.join(temp_dir, "canvas_output")
                os.makedirs(output_folder, exist_ok=True)
                
                # Process images
                processed_count = 0
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                all_files = []
                for root, dirs, files in os.walk(input_folder):
                    for file in files:
                        ext = os.path.splitext(file)[1].lower()
                        if ext in ALLOWED_IMAGE_EXTENSIONS:
                            all_files.append(os.path.join(root, file))
                
                for i, file_path in enumerate(all_files):
                    status_text.text(f"Processing {os.path.basename(file_path)}... ({i+1}/{len(all_files)})")
                    progress_bar.progress((i + 1) / len(all_files))
                    
                    try:
                        with PILImage.open(file_path) as img:
                            width, height = img.size
                            
                            if width < canvas_size and height < canvas_size:
                                # Create transparent canvas
                                canvas = PILImage.new('RGBA', (canvas_size, canvas_size), (255, 255, 255, 0))
                                
                                # Convert image to RGBA if needed
                                if img.mode != 'RGBA':
                                    img = img.convert('RGBA')
                                
                                # Center image on canvas
                                paste_x = (canvas_size - width) // 2
                                paste_y = (canvas_size - height) // 2
                                canvas.paste(img, (paste_x, paste_y), img)
                                
                                # Save processed image
                                new_filename = os.path.splitext(os.path.basename(file_path))[0] + ".png"
                                output_path = os.path.join(output_folder, new_filename)
                                canvas.save(output_path)
                                processed_count += 1
                    except Exception as e:
                        st.warning(f"Error processing {os.path.basename(file_path)}: {str(e)}")
                
                status_text.empty()
                progress_bar.empty()
                
                if processed_count > 0:
                    st.success(f"✅ Successfully processed {processed_count} images")
                    
                    # Create zip file with processed images
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                        for root, dirs, files in os.walk(output_folder):
                            for file in files:
                                file_path = os.path.join(root, file)
                                rel_path = os.path.relpath(file_path, output_folder)
                                zip_file.write(file_path, rel_path)
                    
                    zip_buffer.seek(0)
                    
                    # Download button
                    st.download_button(
                        label="📥 Download Processed Images (ZIP)",
                        data=zip_buffer.getvalue(),
                        file_name="centered_images.zip",
                        mime="application/zip"
                    )
                    
                    # Show preview
                    with st.expander("🖼️ Preview Processed Images"):
                        sample_files = []
                        for root, dirs, files in os.walk(output_folder):
                            for file in files[:3]:  # Get first 3 files from each folder
                                if len(sample_files) < 6:  # Max 6 samples
                                    sample_files.append(os.path.join(root, file))
                        
                        cols = st.columns(3)
                        for i, file_path in enumerate(sample_files):
                            with cols[i % 3]:
                                img = PILImage.open(file_path)
                                st.image(img, caption=os.path.basename(file_path), use_column_width=True)
                        
                        if processed_count > 6:
                            st.info(f"Showing sample images. Total processed: {processed_count}")
                else:
                    st.warning("No images were processed (all images were already larger than the canvas size).")
            
            except Exception as e:
                st.error(f"An error occurred: {str(e)}")
    
    else:
        st.info("👆 Please upload a zip file with images to process.")
        
        # Instructions
        with st.expander("📖 Instructions"):
            st.markdown("""
            **How to use this tool:**
            
            1. Upload a zip file containing images
            2. Set the canvas size (default is 500x500 pixels)
            3. Click "Process Images"
            4. Download the processed images
            
            **Features:**
            - Centers images smaller than canvas size on transparent background
            - Images larger than canvas size are skipped
            - Outputs PNG files to preserve transparency
            - Preserves folder structure
            
            **Note:**
            - Only images smaller than the specified canvas size will be processed
            - Original aspect ratio is maintained
            """)

def extract_zip_to_temp(uploaded_file):
    """Extract uploaded zip file to temporary directory"""
    temp_dir = tempfile.mkdtemp()
    with zipfile.ZipFile(uploaded_file, 'r') as zip_ref:
        zip_ref.extractall(temp_dir)
    return temp_dir

def extract_brand(file_stem):
    """Extract brand name from filename using pattern matching"""
    parts = re.split(r'[ _\-]', file_stem)
    return re.sub(r'[^a-z0-9]', '', parts[0].lower()) if parts else "unknown"

def clean_letters_only(text):
    return re.sub(r'[^A-Za-z]', '', text)

def get_cleaned_filename_without_brand(filename, brand):
    full_letters = clean_letters_only(Path(filename).stem.lower())
    return full_letters.replace(brand, '', 1)

def generate_excel_report(output_folder: Path, marken_index: dict, file_to_factorgroup: dict):
    final_excel_path = output_folder / "IcAt_Overview_Final.xlsx"

    nummer_zu_marke = {v: k for k, v in marken_index.items()}
    data = []
    for file in sorted(output_folder.iterdir()):
        if file.is_file() and file.suffix.lower() in {'.png', '.txt', '.csv', '.md'}:
            name = file.stem
            match = re.match(r"(\d{2})B(\d{2})([a-z0-9]+)", name, re.IGNORECASE)
            if match:
                markennummer, blocknummer, marke = match.groups()
                factor = nummer_zu_marke.get(markennummer, marke)
                
                # Get factorgroup and remove underscores
                factorgroup = file_to_factorgroup.get(file.name, f"{blocknummer}Unknown")
                factorgroup = factorgroup.replace('_', '')  # Remove all underscores
                
                data.append({
                    "factor": factor,
                    "factorgroup": factorgroup,
                    "ID": name
                })

    df_assets = pd.DataFrame(data, columns=["factor", "factorgroup", "ID"])

    reordered_data = []
    for _, row in df_assets.iterrows():
        raw_fg = str(row["ID"])
        group_prefix = raw_fg[:2]
        try:
            group = str(int(group_prefix))
        except ValueError:
            group = group_prefix

        # Also remove underscores from the reordered data
        clean_factor = str(row["factorgroup"]).replace('_', '')
        clean_factorgroup = str(row["factor"]).replace('_', '')

        reordered_data.append({
            "Group": group,
            "factor": clean_factor,
            "factorgroup": clean_factorgroup,
            "ID": row["ID"]
        })

    df_reordered = pd.DataFrame(reordered_data, columns=["Group", "factor", "factorgroup", "ID"])
    df_reordered = df_reordered.sort_values(by='Group', ascending=True).reset_index(drop=True)
    with pd.ExcelWriter(final_excel_path, engine='openpyxl') as writer:
        df_assets.to_excel(writer, index=False, sheet_name="Assets")
        df_reordered.to_excel(writer, index=False, sheet_name="Reordered")

    return final_excel_path


def get_asset_cell(file_path, filename, col_count):
    """Create optimized asset cell for PDF with proper transparency handling"""
    ext = Path(file_path).suffix.lower()
    styles = getSampleStyleSheet()
    
    try:
        if ext in ALLOWED_IMAGE_EXTENSIONS:
            with PILImage.open(file_path) as img:
                # Handle transparency properly
                if img.mode in ('RGBA', 'LA') or (img.mode == 'P' and 'transparency' in img.info):
                    # Create a white background for transparent images
                    background = PILImage.new('RGB', img.size, (255, 255, 255))
                    if img.mode == 'RGBA':
                        # Use alpha channel as mask
                        background.paste(img, mask=img.split()[-1])
                    elif img.mode == 'LA':
                        # Luminance + Alpha
                        img_rgba = img.convert('RGBA')
                        background.paste(img_rgba, mask=img_rgba.split()[-1])
                    elif img.mode == 'P' and 'transparency' in img.info:
                        # Palette mode with transparency
                        img_rgba = img.convert('RGBA')
                        background.paste(img_rgba, mask=img_rgba.split()[-1])
                    img = background
                else:
                    # No transparency, safe to convert to RGB
                    img = img.convert('RGB')
                
                orig_width, orig_height = img.size
                available_width = A4[0] - 40
                max_width = available_width / col_count
                max_height = 120

                aspect_ratio = orig_width / orig_height
                width = max_width
                height = width / aspect_ratio

                if height > max_height:
                    height = max_height
                    width = height * aspect_ratio

                buffer = io.BytesIO()
                # Save as PNG to preserve quality
                img.save(buffer, format='PNG')
                buffer.seek(0)
                
                rl_img = RLImage(buffer, width=width, height=height)

                frame = KeepInFrame(max_width, max_height + 30, content=[
                    rl_img,
                    Paragraph(filename, styles['Normal'])
                ], hAlign='CENTER')

                return frame
        else:
            return Paragraph(f"{filename}", styles['Normal'])
    except Exception as e:
        return Paragraph(f"[Image Error]<br/>{filename}", styles['Normal'])
    

def analyze_files_by_filename(input_folder):
    """Analyze files by extracting brand names from filenames"""
    dateien = []
    marken_set = set()
    renamed_files_by_folder_and_marke = defaultdict(lambda: defaultdict(list))

    for subfolder in sorted(Path(input_folder).iterdir()):
        if not subfolder.is_dir():
            continue

        for file in sorted(subfolder.iterdir()):
            if file.suffix.lower() not in ALL_ALLOWED_EXTENSIONS:
                continue
            marke = extract_brand(file.stem)
            marken_set.add(marke)
            dateien.append({
                "original_file": file,
                "original_folder": subfolder.name,
                "marke": marke
            })

    for eintrag in dateien:
        orig = eintrag["original_file"]
        marke = eintrag["marke"]
        folder_name = eintrag["original_folder"]
        renamed_files_by_folder_and_marke[folder_name][marke].append((orig, orig.name))

    return marken_set, renamed_files_by_folder_and_marke, dateien

def generate_filename_based_pdf_report(input_folder, erste_marke=None, renamed_files_mapping=None):
    """Generate PDF report based on filename brand analysis with correct ID labels"""
    marken_set, renamed_files_by_folder_and_marke, all_files = analyze_files_by_filename(input_folder)
    
    if not marken_set:
        return None, "No brands found in filenames."

    # Create brand index
    marken_index = {}
    if erste_marke and erste_marke in marken_set:
        marken_index[erste_marke] = "01"
        aktuelle_nummer = 2
        for marke in sorted(marken_set):
            if marke != erste_marke:
                marken_index[marke] = f"{aktuelle_nummer:02d}"
                aktuelle_nummer += 1
    else:
        for i, marke in enumerate(sorted(marken_set), 1):
            marken_index[marke] = f"{i:02d}"

    # Generate PDF
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, leftMargin=20, rightMargin=20, topMargin=40, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    nummer_zu_marke = {v: k for k, v in marken_index.items()}
    marken_spalten = sorted(nummer_zu_marke.items())

    for folder in sorted(renamed_files_by_folder_and_marke):
        elements.append(Paragraph(f"<b>Folder: {folder}</b>", styles['Heading2']))

        abschnitt = renamed_files_by_folder_and_marke[folder]
        total = 0
        lines = []
        for nummer, marke in marken_spalten:
            count = len(abschnitt.get(marke, []))
            total += count
            lines.append(f"{nummer} ({marke}): {count}")
        lines.append(f"Total: {total}")
        elements.append(Paragraph("<br/>".join(lines), styles['Normal']))
        elements.append(Spacer(1, 10))

        headers = [f"{nummer} ({marke})" for nummer, marke in marken_spalten]
        data = [headers]
        col_data = []
        max_rows = 0
        for _, marke in marken_spalten:
            eintraege = abschnitt.get(marke, [])
            zellen = []
            for p, original_name in eintraege:
                # Generate the ID name in the same format as used in Excel
                blocknummer = re.sub(r'\D', '', folder)[:2].zfill(2)
                markennummer = marken_index[marke]
                cleaned = get_cleaned_filename_without_brand(original_name, marke)
                # Remove file extension for the ID
                id_name = f"{markennummer}B{blocknummer}{marke}{cleaned}"
                
                # Use the ID name as the label instead of original filename
                cell = get_asset_cell(p, id_name, len(headers))
                zellen.append(cell)
            col_data.append(zellen)
            max_rows = max(max_rows, len(zellen))

        for i in range(max_rows):
            row = []
            for col in col_data:
                row.append(col[i] if i < len(col) else "")
            data.append(row)

        col_width = (A4[0] - 40) / len(headers)
        t = Table(data, colWidths=[col_width] * len(headers))
        t.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(t)
        elements.append(PageBreak())

    # Total overview
    elements.append(Paragraph("<b>Total Overview</b>", styles['Heading2']))
    global_counts = defaultdict(int)
    for folder_data in renamed_files_by_folder_and_marke.values():
        for marke, daten in folder_data.items():
            global_counts[marke] += len(daten)
    
    gesamt = 0
    summary = []
    for nummer, marke in marken_spalten:
        count = global_counts[marke]
        summary.append(f"{marke}: {count} assets")
        gesamt += count
    summary.append(f"Total number of all assets: {gesamt}")
    elements.append(Paragraph("<br/>".join(summary), styles['Normal']))

    # Add brand pages with corrected ID labels
    add_brand_pages_with_id_labels(elements, marken_spalten, renamed_files_by_folder_and_marke, marken_index)

    try:
        doc.build(elements)
        pdf_buffer.seek(0)
        return pdf_buffer, None
    except Exception as e:
        return None, f"Error generating PDF: {str(e)}"

def add_brand_pages_with_id_labels(elements, marken_spalten, renamed_files_by_folder_and_marke, marken_index):
    """Add brand overview pages with ID labels instead of filenames"""
    styles = getSampleStyleSheet()
    for nummer, marke in marken_spalten:
        assets = []
        for folder in renamed_files_by_folder_and_marke:
            folder_assets = renamed_files_by_folder_and_marke[folder].get(marke, [])
            for pfad, original_name in folder_assets:
                # Generate the ID name in the same format as used in Excel
                blocknummer = re.sub(r'\D', '', folder)[:2].zfill(2)
                markennummer = marken_index[marke]
                cleaned = get_cleaned_filename_without_brand(original_name, marke)
                id_name = f"{markennummer}B{blocknummer}{marke}{cleaned}"
                assets.append((pfad, id_name))

        if not assets:
            continue

        elements.append(PageBreak())
        elements.append(Paragraph(f"<b>Brand Overview: {nummer} – {marke}</b>", styles['Heading2']))
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(f"Number of assets: {len(assets)}", styles['Normal']))
        elements.append(Spacer(1, 10))

        headers = ["Asset"] * 4
        data = [headers]
        row = []
        for i, (pfad, id_name) in enumerate(assets):
            # Use ID name as the label
            cell = get_asset_cell(pfad, id_name, 4)
            row.append(cell)
            if len(row) == 4:
                data.append(row)
                row = []
        if row:
            row.extend([""] * (4 - len(row)))
            data.append(row)

        table = Table(data, colWidths=(A4[0] - 40) / 4)
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(table)

def process_files(input_folder, marken_index, output_root):
    """Process and rename files according to brand numbering with sequential counting per brand/block"""
    dateien = []
    renamed_files_by_folder_and_marke = defaultdict(lambda: defaultdict(list))
    file_to_factorgroup = {}

    for subfolder in sorted(Path(input_folder).iterdir()):
        if not subfolder.is_dir():
            continue
        blocknummer = re.sub(r'\D', '', subfolder.name)[:2].zfill(2)
        
        # Count files per brand in this folder for sequential numbering
        brand_counters = defaultdict(int)
        
        # First pass: count files per brand to establish proper numbering
        files_to_process = []
        for file in sorted(subfolder.iterdir()):
            if file.suffix.lower() not in ALL_ALLOWED_EXTENSIONS:
                continue
            marke = extract_brand(file.stem)
            brand_counters[marke] += 1
            files_to_process.append((file, marke))
        
        # Reset counters for actual processing
        brand_counters = defaultdict(int)
        
        # Second pass: process files with sequential numbering
        for file, marke in files_to_process:
            brand_counters[marke] += 1  # Increment counter for this brand
            count_str = f"{brand_counters[marke]:02d}"  # Format as 01, 02, 03, etc.
            
            markennummer = marken_index[marke]
            cleaned = get_cleaned_filename_without_brand(file.name, marke)
            
            # New format: markennummer + B + blocknummer + marke + count + cleaned + extension
            neuer_name = f"{markennummer}B{blocknummer}{marke}{count_str}{cleaned}{file.suffix.lower()}"
            ziel = output_root / neuer_name

            if file.suffix.lower() in ALLOWED_IMAGE_EXTENSIONS:
                with PILImage.open(file) as img:
                    # Preserve transparency instead of converting to RGB
                    if img.mode in ('RGBA', 'LA') or (img.mode == 'P' and 'transparency' in img.info):
                        # If image already has transparency, preserve it
                        img.save(ziel, format='PNG')
                    else:
                        # For images without transparency, convert to RGBA to allow transparency
                        img.convert("RGBA").save(ziel, format='PNG')
            else:
                shutil.copy2(file, ziel)

            renamed_files_by_folder_and_marke[subfolder.name][marke].append((ziel, neuer_name))
            file_to_factorgroup[neuer_name] = subfolder.name

    return renamed_files_by_folder_and_marke, file_to_factorgroup

def brand_renamer_tool():
    st.header("Advanced Brand File Processor")
    st.markdown("Automatically rename and organize brand assets with numbering and generate reports.")
   
    # Initialize session state for persistent results
    if 'processing_complete' not in st.session_state:
        st.session_state.processing_complete = False
    if 'processed_data' not in st.session_state:
        st.session_state.processed_data = {}
   
    uploaded_file = st.file_uploader(
        "Choose a ZIP file",
        type=['zip'],
    )
   
    # Reset processing state when new file is uploaded
    if uploaded_file and 'uploaded_file_name' in st.session_state:
        if st.session_state.uploaded_file_name != uploaded_file.name:
            st.session_state.processing_complete = False
            st.session_state.processed_data = {}
   
    if uploaded_file:
        # Store uploaded file name to detect changes
        st.session_state.uploaded_file_name = uploaded_file.name
       
        if not st.session_state.processing_complete:
            with st.spinner("Extracting and analyzing files..."):
                temp_dir = extract_zip_to_temp(uploaded_file)
                input_folder = Path(temp_dir)
               
                items = os.listdir(temp_dir)
                if len(items) == 1 and os.path.isdir(os.path.join(temp_dir, items[0])):
                    input_folder = Path(temp_dir) / items[0]
               
                output_folder = Path(temp_dir) / "output"
                output_folder.mkdir(exist_ok=True)
               
                marken_set, _, _ = analyze_files_by_filename(input_folder)
               
                if not marken_set:
                    st.error("No brands found in the uploaded files.")
                    return
               
                st.success(f"Detected brands: {', '.join(sorted(marken_set))}")
               
                # Brand numbering selection
                st.subheader("Brand Numbering")
                erste_marke = st.selectbox(
                    "Which brand should be number 01?",
                    options=sorted(marken_set),
                    index=0
                )
               
                marken_index = {erste_marke: "01"}
                aktuelle_nummer = 2
                for marke in sorted(marken_set):
                    if marke != erste_marke:
                        marken_index[marke] = f"{aktuelle_nummer:02d}"
                        aktuelle_nummer += 1
               
                if st.button("Process Files and Generate Reports", type="primary"):
                    with st.spinner("Processing files..."):
                        # Process and rename files
                        renamed_files, file_to_factorgroup = process_files(
                            input_folder,
                            marken_index,
                            output_folder
                        )
                       
                        # Also rename actual files in output folder to have .png extensions
                        for file_path in output_folder.iterdir():
                            if file_path.is_file() and not file_path.suffix:
                                # File has no extension, add .png
                                new_file_path = file_path.with_suffix('.png')
                                try:
                                    file_path.rename(new_file_path)
                                except Exception as e:
                                    print(f"Could not rename {file_path}: {e}")
                       
                        # Generate PDF with extensions - use the modified function
                        pdf_buffer, error = generate_filename_based_pdf_report_with_extensions(
                            input_folder,
                            erste_marke,
                            marken_index  # Pass marken_index for ID generation
                        )
                        if error:
                            st.error(f"PDF generation failed: {error}")
                       
                        excel_path = generate_excel_report(output_folder, marken_index, file_to_factorgroup)
                       
                        # Create individual file buffers
                        zip_buffer = io.BytesIO()
                        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                            for file in output_folder.iterdir():
                                zip_file.write(file, file.name)
                        zip_buffer.seek(0)
                       
                        # Create combined download with all files
                        combined_zip_buffer = io.BytesIO()
                        with zipfile.ZipFile(combined_zip_buffer, 'w', zipfile.ZIP_DEFLATED) as combined_zip:
                            # Add processed files
                            for file in output_folder.iterdir():
                                combined_zip.write(file, f"processed_files/{file.name}")
                           
                            # Add PDF report if available
                            if pdf_buffer:
                                combined_zip.writestr("reports/Brand_Assets_Report.pdf", pdf_buffer.getvalue())
                           
                            # Add Excel report
                            with open(excel_path, 'rb') as excel_file:
                                combined_zip.writestr("reports/Brand_Assets_Overview.xlsx", excel_file.read())
                       
                        combined_zip_buffer.seek(0)
                       
                        # Store processed data in session state
                        st.session_state.processed_data = {
                            'pdf_buffer': pdf_buffer,
                            'excel_path': excel_path,
                            'zip_buffer': zip_buffer,
                            'combined_zip_buffer': combined_zip_buffer,
                            'temp_dir': temp_dir,
                            'renamed_files': renamed_files
                        }
                        st.session_state.processing_complete = True
                       
                        # Force rerun to show download buttons
                        st.rerun()
       
        # Show download options if processing is complete
        if st.session_state.processing_complete:
            st.success("Processing complete!")
           
            # Display a preview of processed names with extensions
            with st.expander("Preview Processed Names"):
                if 'renamed_files' in st.session_state.processed_data:
                    st.write("Sample processed filenames (with .png extensions in PDF):")
                    # Handle the case where renamed_files might be a defaultdict
                    renamed_files_data = st.session_state.processed_data['renamed_files']
                    
                    # Extract sample file names safely
                    sample_files = []
                    if hasattr(renamed_files_data, 'items'):
                        # It's a dictionary-like object
                        for folder_name, brand_data in list(renamed_files_data.items())[:3]:  # First 3 folders
                            for brand_name, files in list(brand_data.items())[:2]:  # First 2 brands per folder
                                for file_path, new_name in files[:2]:  # First 2 files per brand
                                    if isinstance(new_name, str):
                                        pdf_name = Path(new_name).stem + '.png'
                                        sample_files.append((Path(file_path).name, pdf_name))
                                    if len(sample_files) >= 10:
                                        break
                                if len(sample_files) >= 10:
                                    break
                            if len(sample_files) >= 10:
                                break
                    
                    # Display the sample files
                    for old_name, pdf_name in sample_files:
                        st.write(f"{old_name} → {pdf_name}")
           
            # Option to start over
            if st.button("🔄 Process New File", help="Clear results and upload a new file"):
                st.session_state.processing_complete = False
                st.session_state.processed_data = {}
                if 'temp_dir' in st.session_state.processed_data:
                    try:
                        shutil.rmtree(st.session_state.processed_data['temp_dir'])
                    except:
                        pass
                st.rerun()
           
            st.markdown("---")
           
            # Combined download option (recommended)
            st.subheader("📦 Complete Package Download")
            st.markdown("**Recommended:** Download everything in one convenient package")
           
            if st.session_state.processed_data.get('combined_zip_buffer'):
                st.download_button(
                    label="🎁 Download Complete Package (All Files + Reports)",
                    data=st.session_state.processed_data['combined_zip_buffer'].getvalue(),
                    file_name="Brand_Assets_Complete_Package.zip",
                    mime="application/zip",
                    type="primary"
                )
           
            st.markdown("---")
           
            # Individual download options
            st.subheader("📁 Individual Downloads")
            st.markdown("Or download items separately:")
           
            col1, col2, col3 = st.columns(3)
           
            with col1:
                if st.session_state.processed_data.get('pdf_buffer'):
                    st.download_button(
                        label="📄 PDF Report",
                        data=st.session_state.processed_data['pdf_buffer'].getvalue(),
                        file_name="Brand_Assets_Report.pdf",
                        mime="application/pdf"
                    )
                else:
                    st.info("PDF report not available")
           
            with col2:
                if st.session_state.processed_data.get('excel_path'):
                    try:
                        with open(st.session_state.processed_data['excel_path'], 'rb') as excel_file:
                            st.download_button(
                                label="📊 Excel Report",
                                data=excel_file.read(),
                                file_name="Brand_Assets_Overview.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    except:
                        st.info("Excel report not available")
           
            with col3:
                if st.session_state.processed_data.get('zip_buffer'):
                    st.download_button(
                        label="📦 Processed Files",
                        data=st.session_state.processed_data['zip_buffer'].getvalue(),
                        file_name="Brand_Assets_Processed.zip",
                        mime="application/zip"
                    )
           
            st.markdown("---")
            st.info("💡 **Tip:** Downloads will remain available until you upload a new file or click 'Process New File'")
   
    else:
        # Reset session state when no file is uploaded
        if st.session_state.processing_complete:
            st.session_state.processing_complete = False
            st.session_state.processed_data = {}
       
        st.info("👆 Please upload a zip file to get started.")
       
        with st.expander("📖 Instructions"):
            st.markdown("""
            Upload a ZIP file containing brand asset folders. The tool will:
            1. Detect all brands automatically
            2. Let you choose which brand gets number 01
            3. Process and rename all files with proper numbering
            4. Generate a PDF report showing all assets with filenames including .png extensions
            5. Create an Excel overview of the processed files
            6. Package everything for easy download
            """)


def generate_filename_based_pdf_report_with_extensions(input_folder, erste_marke=None, marken_index=None):
    """Generate PDF report with .png extensions in image labels"""
    marken_set, renamed_files_by_folder_and_marke, all_files = analyze_files_by_filename(input_folder)
    
    if not marken_set:
        return None, "No brands found in filenames."

    # Use provided marken_index or create new one
    if not marken_index:
        marken_index = {}
        if erste_marke and erste_marke in marken_set:
            marken_index[erste_marke] = "01"
            aktuelle_nummer = 2
            for marke in sorted(marken_set):
                if marke != erste_marke:
                    marken_index[marke] = f"{aktuelle_nummer:02d}"
                    aktuelle_nummer += 1
        else:
            for i, marke in enumerate(sorted(marken_set), 1):
                marken_index[marke] = f"{i:02d}"

    # Generate PDF
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, leftMargin=20, rightMargin=20, topMargin=40, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    nummer_zu_marke = {v: k for k, v in marken_index.items()}
    marken_spalten = sorted(nummer_zu_marke.items())

    for folder in sorted(renamed_files_by_folder_and_marke):
        elements.append(Paragraph(f"<b>Folder: {folder}</b>", styles['Heading2']))

        abschnitt = renamed_files_by_folder_and_marke[folder]
        total = 0
        lines = []
        for nummer, marke in marken_spalten:
            count = len(abschnitt.get(marke, []))
            total += count
            lines.append(f"{nummer} ({marke}): {count}")
        lines.append(f"Total: {total}")
        elements.append(Paragraph("<br/>".join(lines), styles['Normal']))
        elements.append(Spacer(1, 10))

        headers = [f"{nummer} ({marke})" for nummer, marke in marken_spalten]
        data = [headers]
        col_data = []
        max_rows = 0
        for _, marke in marken_spalten:
            eintraege = abschnitt.get(marke, [])
            zellen = []
            for p, original_name in eintraege:
                # Generate the ID name and add .png extension for display
                blocknummer = re.sub(r'\D', '', folder)[:2].zfill(2)
                markennummer = marken_index[marke]
                cleaned = get_cleaned_filename_without_brand(original_name, marke)
                original_ext = Path(original_name).suffix.lower()  # Get actual extension
                id_name = f"{markennummer}B{blocknummer}{marke}{cleaned}{original_ext}"  # Add .png here
                
                cell = get_asset_cell(p, id_name, len(headers))
                zellen.append(cell)
            col_data.append(zellen)
            max_rows = max(max_rows, len(zellen))

        for i in range(max_rows):
            row = []
            for col in col_data:
                row.append(col[i] if i < len(col) else "")
            data.append(row)

        col_width = (A4[0] - 40) / len(headers)
        t = Table(data, colWidths=[col_width] * len(headers))
        t.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(t)
        elements.append(PageBreak())

    # Total overview
    elements.append(Paragraph("<b>Total Overview</b>", styles['Heading2']))
    global_counts = defaultdict(int)
    for folder_data in renamed_files_by_folder_and_marke.values():
        for marke, daten in folder_data.items():
            global_counts[marke] += len(daten)
    
    gesamt = 0
    summary = []
    for nummer, marke in marken_spalten:
        count = global_counts[marke]
        summary.append(f"{marke}: {count} assets")
        gesamt += count
    summary.append(f"Total number of all assets: {gesamt}")
    elements.append(Paragraph("<br/>".join(summary), styles['Normal']))

    # Add brand pages with .png extensions
    add_brand_pages_with_png_extensions(elements, marken_spalten, renamed_files_by_folder_and_marke, marken_index)

    try:
        doc.build(elements)
        pdf_buffer.seek(0)
        return pdf_buffer, None
    except Exception as e:
        return None, f"Error generating PDF: {str(e)}"

def add_brand_pages_with_png_extensions(elements, marken_spalten, renamed_files_by_folder_and_marke, marken_index):
    """Add brand overview pages with .png extensions in labels"""
    styles = getSampleStyleSheet()
    for nummer, marke in marken_spalten:
        assets = []
        for folder in renamed_files_by_folder_and_marke:
            folder_assets = renamed_files_by_folder_and_marke[folder].get(marke, [])
            for pfad, original_name in folder_assets:
                # Generate the ID name with .png extension for display
                blocknummer = re.sub(r'\D', '', folder)[:2].zfill(2)
                markennummer = marken_index[marke]
                cleaned = get_cleaned_filename_without_brand(original_name, marke)
                original_ext = Path(original_name).suffix.lower()  # Get actual extension
                id_name = f"{markennummer}B{blocknummer}{marke}{cleaned}{original_ext}"  # Add .png here
                assets.append((pfad, id_name))

        if not assets:
            continue

        elements.append(PageBreak())
        elements.append(Paragraph(f"<b>Brand Overview: {nummer} – {marke}</b>", styles['Heading2']))
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(f"Number of assets: {len(assets)}", styles['Normal']))
        elements.append(Spacer(1, 10))

        headers = ["Asset"] * 4
        data = [headers]
        row = []
        for i, (pfad, id_name) in enumerate(assets):
            # Use ID name with .png extension as the label
            cell = get_asset_cell(pfad, id_name, 4)
            row.append(cell)
            if len(row) == 4:
                data.append(row)
                row = []
        if row:
            row.extend([""] * (4 - len(row)))
            data.append(row)

        table = Table(data, colWidths=(A4[0] - 40) / 4)
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(table)


def main():
    st.title("🔧 Brand Assets Tools")

    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
        "Extract Images from Excel", 
        "Name assets by brand",
        "Canvas white to Transparent",
        "Asset Overview", 
        "Name stims by block + create output files​",
        "Resize",
        "Place on Canvas"

    ])

    with tab1:
         excel_extractor_tool()
    with tab2:
         file_renamer_tool()  
    with tab3:
        white_to_transparent_tool()  
    with tab4:
         pdf_generator_tool() 
    with tab5:
         brand_renamer_tool() 
    with tab6:
        resize_with_transparent_canvas_tool()
    with tab7:
        center_on_canvas_tool()

if __name__ == "__main__":
    main()