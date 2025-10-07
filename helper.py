import io
import os
import re
import shutil
import tempfile
import zipfile
from collections import defaultdict
from pathlib import Path
import pandas as pd
import streamlit as st
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Image as RLImage, 
    Paragraph, Spacer, PageBreak, KeepInFrame
)
import openpyxl
from openpyxl.styles import PatternFill



ALLOWED_IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.tif', '.webp', '.svg', '.ico', '.heic', '.heif', '.raw'}
ALLOWED_TEXT_EXTENSIONS = {'.txt', '.md', '.csv', '.doc', '.docx', '.pdf', '.rtf', '.odt', '.tex'}
ALLOWED_VIDEO_EXTENSIONS = {'.mp4', '.avi', '.mov', '.wmv', '.flv', '.mkv', '.webm', '.m4v', '.3gp', '.ogv', '.mpg', '.mpeg', '.vob', '.mts', '.m2ts'}
ALLOWED_AUDIO_EXTENSIONS = {'.mp3', '.wav', '.aac', '.flac', '.ogg', '.m4a', '.wma', '.opus', '.aiff', '.ape', '.alac'}
ALLOWED_ARCHIVE_EXTENSIONS = {'.zip', '.rar', '.7z', '.tar', '.gz', '.bz2', '.xz'}
ALLOWED_CODE_EXTENSIONS = {'.py', '.js', '.html', '.css', '.java', '.cpp', '.c', '.h', '.php', '.rb', '.go', '.rs', '.swift', '.kt'}
ALLOWED_DATA_EXTENSIONS = {'.json', '.xml', '.yaml', '.yml', '.sql', '.db', '.sqlite', '.xlsx', '.xls'}
ALLOWED_OTHER_EXTENSIONS = {'.psd', '.ai', '.sketch', '.fig', '.eps', '.indd', '.dwg', '.stl', '.obj', '.fbx', '.blend'}

# Combine all extensions - if a file has ANY extension or NO extension, process it
ALL_ALLOWED_EXTENSIONS = (ALLOWED_IMAGE_EXTENSIONS | ALLOWED_TEXT_EXTENSIONS | 
                          ALLOWED_VIDEO_EXTENSIONS | ALLOWED_AUDIO_EXTENSIONS |
                          ALLOWED_ARCHIVE_EXTENSIONS | ALLOWED_CODE_EXTENSIONS |
                          ALLOWED_DATA_EXTENSIONS | ALLOWED_OTHER_EXTENSIONS)

# Helper functions from both files
def extract_zip_to_temp(uploaded_file):
    """Extract uploaded zip file to temporary directory"""
    temp_dir = tempfile.mkdtemp()
    with zipfile.ZipFile(uploaded_file, 'r') as zip_ref:
        zip_ref.extractall(temp_dir)
    return temp_dir

def extract_images_from_xlsx_zip_method(xlsx_path, output_dir, target_sheet=None):
    """
    Fallback method: Extract all images from xlsx using zip method
    """
    try:
        extracted_files = []
        image_count = 0
        
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            # Get available sheets
            available_sheets = []
            try:
                with zip_ref.open('xl/workbook.xml') as f:
                    content = f.read().decode('utf-8')
                sheet_matches = re.findall(r'<sheet name="([^"]+)"', content)
                available_sheets = [match for match in sheet_matches]
            except:
                available_sheets = ["Sheet1"]
            
            # Find all media files
            media_files = [f for f in zip_ref.namelist() if f.startswith('xl/media/')]
            
            if not media_files:
                return 0, [], "No images found in Excel file", available_sheets
            
            # Extract all images (since precise sheet mapping is complex)
            for media_file in media_files:
                try:
                    image_data = zip_ref.read(media_file)
                    original_filename = os.path.basename(media_file)
                    
                    # Create output filename
                    if target_sheet and target_sheet != "All worksheets":
                        output_filename = f"{target_sheet}_{original_filename}"
                    else:
                        output_filename = original_filename
                    
                    output_path = os.path.join(output_dir, output_filename)
                    
                    with open(output_path, 'wb') as f:
                        f.write(image_data)
                    
                    extracted_files.append(output_path)
                    image_count += 1
                    
                except Exception as e:
                    print(f"Error extracting {media_file}: {e}")
                    continue
        
        warning_msg = None
        if target_sheet and target_sheet != "All worksheets":
            warning_msg = f"Note: Extracted all images from file. Sheet-specific extraction not fully supported with this method."
        
        return image_count, extracted_files, warning_msg, available_sheets
        
    except Exception as e:
        return 0, [], f"Error processing Excel file: {str(e)}", []
    
def process_files(input_folder, marken_index, output_root):
    """Process and rename files according to brand numbering with sequential counting per brand/block"""
    dateien = []
    renamed_files_by_folder_and_marke = defaultdict(lambda: defaultdict(list))
    file_to_factorgroup = {}

    for subfolder in sorted(Path(input_folder).iterdir()):
        if not subfolder.is_dir():
            continue
        blocknummer = re.sub(r'\D', '', subfolder.name)[:2].zfill(2)
        
        # Count files per brand in this folder for sequential numbering
        brand_counters = defaultdict(int)
        
        # First pass: count files per brand to establish proper numbering
        files_to_process = []
        for file in sorted(subfolder.iterdir()):
            if file.suffix.lower() not in ALL_ALLOWED_EXTENSIONS:
                continue
            marke = extract_brand(file.stem)
            brand_counters[marke] += 1
            files_to_process.append((file, marke))
        
        # Reset counters for actual processing
        brand_counters = defaultdict(int)
        
        # Second pass: process files with sequential numbering
        for file, marke in files_to_process:
            brand_counters[marke] += 1  # Increment counter for this brand
            count_str = f"{brand_counters[marke]:02d}"  # Format as 01, 02, 03, etc.
            
            markennummer = marken_index[marke]
            cleaned = get_cleaned_filename_without_brand(file.name, marke)
            
            # New format: markennummer + B + blocknummer + marke + count + cleaned + extension
            neuer_name = f"{markennummer}B{blocknummer}{marke}{count_str}{cleaned}{file.suffix.lower()}"
            ziel = output_root / neuer_name

            if file.suffix.lower() in ALLOWED_IMAGE_EXTENSIONS:
                with PILImage.open(file) as img:
                    # Preserve transparency instead of converting to RGB
                    if img.mode in ('RGBA', 'LA') or (img.mode == 'P' and 'transparency' in img.info):
                        # If image already has transparency, preserve it
                        img.save(ziel, format='PNG')
                    else:
                        # For images without transparency, convert to RGBA to allow transparency
                        img.convert("RGBA").save(ziel, format='PNG')
            else:
                shutil.copy2(file, ziel)

            renamed_files_by_folder_and_marke[subfolder.name][marke].append((ziel, neuer_name))
            file_to_factorgroup[neuer_name] = subfolder.name

    return renamed_files_by_folder_and_marke, file_to_factorgroup

def is_valid_file(file_path):
    """Check if file should be processed - accept ALL files except hidden/system files"""
    filename = Path(file_path).name
    # Skip hidden files and system files
    if filename.startswith('.') or filename.startswith('~') or filename == 'Thumbs.db':
        return False
    # Accept all other files
    return True

def generate_excel_report(output_folder: Path, marken_index: dict, file_to_factorgroup: dict):
    """Generate Excel report with 'language' column, exceptions, and highlighting"""
    final_excel_path = output_folder / "IcAt_Overview_Final.xlsx"

    nummer_zu_marke = {v: k for k, v in marken_index.items()}
    rows = []  # wir führen Flags für Highlighting mit

    for file in sorted(output_folder.iterdir()):
        # Process ALL files, not just specific extensions
        if file.is_file() and is_valid_file(file):
            id_base = file.stem                        # ID ohne Extension
            orig_ext = file.suffix.lower() if file.suffix else ''  # z. B. ".png"

            match = re.match(r"(\d{2})B(\d{2})([a-z0-9]+)", id_base, re.IGNORECASE)
            if not match:
                continue

            markennummer, blocknummer, marke_chunk = match.groups()
            factor = nummer_zu_marke.get(markennummer, marke_chunk)

            # factorgroup aus Mapping, Unterstriche entfernen
            factorgroup = file_to_factorgroup.get(file.name, f"{blocknummer}Unknown")
            factorgroup = str(factorgroup).replace('_', '')

            # Flags (für Regeln & Highlight – basieren auf Originalendung + ID)
            id_upper = id_base.upper()
            has_B20 = ("B20" in id_upper)
            is_txt = (orig_ext == ".txt")
            is_av = (orig_ext in ALLOWED_VIDEO_EXTENSIONS or orig_ext in ALLOWED_AUDIO_EXTENSIONS)
            has_B12_B14_B18 = any(tag in id_upper for tag in ("B12", "B14", "B18"))

            # language gemäß besprochener Logik:
            # A) B12/B14/B18 + .txt => Inhalt der Textdatei 1:1 in language
            if has_B12_B14_B18 and is_txt:
                try:
                    language_val = file.read_text(encoding="utf-8")
                except Exception:
                    language_val = file.read_bytes().decode("utf-8", errors="ignore")
            # B) B20 + Audio/Video => language = ID + ".png"
            elif has_B20 and is_av:
                language_val = f"{id_base}.png"
            # C) Standard => language = ID (+ Original-Endung, falls vorhanden)
            else:
                language_val = f"{id_base}{orig_ext}" if orig_ext else id_base

            rows.append({
                "factor": factor,
                "factorgroup": factorgroup,
                "ID": id_base,            # wie bisher: ohne Extension
                "language": language_val, # neu
                "_hl_txt": is_txt,
                "_hl_av": is_av,
                "_hl_b20": has_B20,
            })

    # DataFrame für Assets
    df_assets = pd.DataFrame(rows)
    if df_assets.empty:
        # trotzdem eine leere Datei mit den richtigen Sheets/Spalten schreiben
        with pd.ExcelWriter(final_excel_path, engine='openpyxl') as writer:
            pd.DataFrame(columns=["factor","factorgroup","ID","language"]).to_excel(writer, index=False, sheet_name="Assets")
            pd.DataFrame(columns=["Group","factor","factorgroup","ID","language"]).to_excel(writer, index=False, sheet_name="Reordered")
        return final_excel_path

    # Sichtbare Spalten
    df_assets_visible = df_assets[["factor", "factorgroup", "ID", "language"]].copy()

    # Reordered (bestehende Logik beibehalten: factor/factorgroup getauscht)
    reordered = []
    for _, r in df_assets.iterrows():
        raw_id = str(r["ID"])
        group_prefix = raw_id[:2]
        try:
            group_val = str(int(group_prefix))  # führende Nullen entfernen
        except ValueError:
            group_val = group_prefix

        clean_factor = str(r["factorgroup"]).replace("_", "")
        clean_factorgroup = str(r["factor"]).replace("_", "")

        reordered.append({
            "Group": group_val,
            "factor": clean_factor,          # wie bisher vertauscht
            "factorgroup": clean_factorgroup,
            "ID": raw_id,
            "language": r["language"],       # mitnehmen
            "_hl_txt": r["_hl_txt"],
            "_hl_av": r["_hl_av"],
            "_hl_b20": r["_hl_b20"],
        })

    df_reordered = pd.DataFrame(reordered).sort_values(by="Group", ascending=True).reset_index(drop=True)
    df_reordered_visible = df_reordered[["Group", "factor", "factorgroup", "ID", "language"]].copy()

    # Schreiben + Highlight
    with pd.ExcelWriter(final_excel_path, engine='openpyxl') as writer:
        df_assets_visible.to_excel(writer, index=False, sheet_name="Assets")
        df_reordered_visible.to_excel(writer, index=False, sheet_name="Reordered")

        wb = writer.book
        ws_assets = writer.sheets["Assets"]
        ws_reordered = writer.sheets["Reordered"]

        # dezentes Gelb für Highlights
        hl_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

        # Spaltenindex der 'language'-Spalte (1-basiert; A=1)
        # Assets: factor | factorgroup | ID | language
        assets_language_col_idx = 4
        # Reordered: Group | factor | factorgroup | ID | language
        reordered_language_col_idx = 5

        # Highlight, wenn: Original .txt ODER Original AV ODER ID enthält B20
        for i, r in enumerate(df_assets.itertuples(index=False), start=2):
            if (r._hl_txt or r._hl_av or r._hl_b20):
                ws_assets.cell(row=i, column=assets_language_col_idx).fill = hl_fill

        for i, r in enumerate(df_reordered.itertuples(index=False), start=2):
            if (r._hl_txt or r._hl_av or r._hl_b20):
                ws_reordered.cell(row=i, column=reordered_language_col_idx).fill = hl_fill

    return final_excel_path


def add_brand_pages_with_png_extensions(elements, marken_spalten, renamed_files_by_folder_and_marke, marken_index):
    """Add brand overview pages with .png extensions in labels"""
    styles = getSampleStyleSheet()
    for nummer, marke in marken_spalten:
        assets = []
        for folder in renamed_files_by_folder_and_marke:
            folder_assets = renamed_files_by_folder_and_marke[folder].get(marke, [])
            for pfad, original_name in folder_assets:
                # Generate the ID name with .png extension for display
                blocknummer = re.sub(r'\D', '', folder)[:2].zfill(2)
                markennummer = marken_index[marke]
                cleaned = get_cleaned_filename_without_brand(original_name, marke)
                original_ext = Path(original_name).suffix.lower()  # Get actual extension
                id_name = f"{markennummer}B{blocknummer}{marke}{cleaned}{original_ext}"  # Add .png here
                assets.append((pfad, id_name))

        if not assets:
            continue

        elements.append(PageBreak())
        elements.append(Paragraph(f"<b>Brand Overview: {nummer} – {marke}</b>", styles['Heading2']))
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(f"Number of assets: {len(assets)}", styles['Normal']))
        elements.append(Spacer(1, 10))

        headers = ["Asset"] * 4
        data = [headers]
        row = []
        for i, (pfad, id_name) in enumerate(assets):
            # Use ID name with .png extension as the label
            cell = get_asset_cell(pfad, id_name, 4)
            row.append(cell)
            if len(row) == 4:
                data.append(row)
                row = []
        if row:
            row.extend([""] * (4 - len(row)))
            data.append(row)

        table = Table(data, colWidths=(A4[0] - 40) / 4)
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(table)


def generate_filename_based_pdf_report_with_extensions(input_folder, erste_marke=None, marken_index=None):
    """Generate PDF report with .png extensions in image labels"""
    marken_set, renamed_files_by_folder_and_marke, all_files = analyze_files_by_filename(input_folder)
    
    if not marken_set:
        return None, "No brands found in filenames."

    # Use provided marken_index or create new one
    if not marken_index:
        marken_index = {}
        if erste_marke and erste_marke in marken_set:
            marken_index[erste_marke] = "01"
            aktuelle_nummer = 2
            for marke in sorted(marken_set):
                if marke != erste_marke:
                    marken_index[marke] = f"{aktuelle_nummer:02d}"
                    aktuelle_nummer += 1
        else:
            for i, marke in enumerate(sorted(marken_set), 1):
                marken_index[marke] = f"{i:02d}"

    # Generate PDF
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, leftMargin=20, rightMargin=20, topMargin=40, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    nummer_zu_marke = {v: k for k, v in marken_index.items()}
    marken_spalten = sorted(nummer_zu_marke.items())

    for folder in sorted(renamed_files_by_folder_and_marke):
        elements.append(Paragraph(f"<b>Folder: {folder}</b>", styles['Heading2']))

        abschnitt = renamed_files_by_folder_and_marke[folder]
        total = 0
        lines = []
        for nummer, marke in marken_spalten:
            count = len(abschnitt.get(marke, []))
            total += count
            lines.append(f"{nummer} ({marke}): {count}")
        lines.append(f"Total: {total}")
        elements.append(Paragraph("<br/>".join(lines), styles['Normal']))
        elements.append(Spacer(1, 10))

        headers = [f"{nummer} ({marke})" for nummer, marke in marken_spalten]
        data = [headers]
        col_data = []
        max_rows = 0
        for _, marke in marken_spalten:
            eintraege = abschnitt.get(marke, [])
            zellen = []
            for p, original_name in eintraege:
                # Generate the ID name and add .png extension for display
                blocknummer = re.sub(r'\D', '', folder)[:2].zfill(2)
                markennummer = marken_index[marke]
                cleaned = get_cleaned_filename_without_brand(original_name, marke)
                original_ext = Path(original_name).suffix.lower()  # Get actual extension
                id_name = f"{markennummer}B{blocknummer}{marke}{cleaned}{original_ext}"  # Add .png here
                
                cell = get_asset_cell(p, id_name, len(headers))
                zellen.append(cell)
            col_data.append(zellen)
            max_rows = max(max_rows, len(zellen))

        for i in range(max_rows):
            row = []
            for col in col_data:
                row.append(col[i] if i < len(col) else "")
            data.append(row)

        col_width = (A4[0] - 40) / len(headers)
        t = Table(data, colWidths=[col_width] * len(headers))
        t.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(t)
        elements.append(PageBreak())

    # Total overview
    elements.append(Paragraph("<b>Total Overview</b>", styles['Heading2']))
    global_counts = defaultdict(int)
    for folder_data in renamed_files_by_folder_and_marke.values():
        for marke, daten in folder_data.items():
            global_counts[marke] += len(daten)
    
    gesamt = 0
    summary = []
    for nummer, marke in marken_spalten:
        count = global_counts[marke]
        summary.append(f"{marke}: {count} assets")
        gesamt += count
    summary.append(f"Total number of all assets: {gesamt}")
    elements.append(Paragraph("<br/>".join(summary), styles['Normal']))

    # Add brand pages with .png extensions
    add_brand_pages_with_png_extensions(elements, marken_spalten, renamed_files_by_folder_and_marke, marken_index)

    try:
        doc.build(elements)
        pdf_buffer.seek(0)
        return pdf_buffer, None
    except Exception as e:
        return None, f"Error generating PDF: {str(e)}"
    

def extract_images_from_xlsx_by_sheet(xlsx_path, output_dir, target_sheet=None):
    """
    Extract images from .xlsx files with reliable sheet-specific extraction
    Uses openpyxl to directly access worksheet images
    """
    try:
        extracted_files = []
        image_count = 0
        
        # Load workbook with openpyxl
        workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
        available_sheets = workbook.sheetnames
        
        # Determine which sheets to process
        if target_sheet is None or target_sheet == "All worksheets":
            sheets_to_process = available_sheets
        else:
            sheets_to_process = [target_sheet] if target_sheet in available_sheets else []
        
        if not sheets_to_process:
            workbook.close()
            return 0, [], "No valid sheets found to extract from", available_sheets
        
        # Process each sheet
        for sheet_name in sheets_to_process:
            try:
                worksheet = workbook[sheet_name]
                
                # Check if worksheet has any images
                if hasattr(worksheet, '_images') and worksheet._images:
                    sheet_image_count = 0
                    for image in worksheet._images:
                        try:
                            # Get image data
                            if hasattr(image, '_data'):
                                image_data = image._data()
                            else:
                                # Try alternative method
                                continue
                            
                            # Determine file extension from image format
                            image_format = 'png'  # default
                            if hasattr(image, 'format') and image.format:
                                image_format = image.format.lower()
                            elif len(image_data) > 4:
                                # Try to detect format from data
                                if image_data[:4] == b'\x89PNG':
                                    image_format = 'png'
                                elif image_data[:3] == b'\xFF\xD8\xFF':
                                    image_format = 'jpg'
                                elif image_data[:6] == b'GIF87a' or image_data[:6] == b'GIF89a':
                                    image_format = 'gif'
                            
                            # Create filename
                            if len(sheets_to_process) > 1:
                                filename = f"{sheet_name}_image_{sheet_image_count + 1:03d}.{image_format}"
                            else:
                                filename = f"image_{sheet_image_count + 1:03d}.{image_format}"
                            
                            # Save image
                            output_path = os.path.join(output_dir, filename)
                            with open(output_path, 'wb') as f:
                                f.write(image_data)
                            
                            extracted_files.append(output_path)
                            sheet_image_count += 1
                            image_count += 1
                            
                        except Exception as e:
                            print(f"Error extracting image from {sheet_name}: {e}")
                            continue
                    
                    print(f"Extracted {sheet_image_count} images from sheet: {sheet_name}")
                
                # Also check for drawing objects (alternative method)
                if hasattr(worksheet, 'drawing') and worksheet.drawing:
                    for drawing in worksheet.drawing:
                        if hasattr(drawing, '_images'):
                            for img in drawing._images:
                                try:
                                    # Similar extraction logic for drawing images
                                    if hasattr(img, '_data'):
                                        image_data = img._data()
                                        image_format = 'png'
                                        
                                        if len(sheets_to_process) > 1:
                                            filename = f"{sheet_name}_drawing_{image_count + 1:03d}.{image_format}"
                                        else:
                                            filename = f"drawing_{image_count + 1:03d}.{image_format}"
                                        
                                        output_path = os.path.join(output_dir, filename)
                                        with open(output_path, 'wb') as f:
                                            f.write(image_data)
                                        
                                        extracted_files.append(output_path)
                                        image_count += 1
                                        
                                except Exception:
                                    continue
                            
            except Exception as e:
                print(f"Error processing sheet {sheet_name}: {e}")
                continue
        
        workbook.close()
        
        # If openpyxl method didn't find images, fall back to zip extraction
        if image_count == 0:
            return extract_images_from_xlsx_zip_method(xlsx_path, output_dir, target_sheet)
        
        return image_count, extracted_files, None, available_sheets
        
    except Exception as e:
        return extract_images_from_xlsx_zip_method(xlsx_path, output_dir, target_sheet)

def extract_images_from_xlsx(xlsx_file, output_dir):
    """Extract images from Excel file to specified directory"""
    count = 0
    extracted_files = []
    
    try:
        with zipfile.ZipFile(xlsx_file, 'r') as z:
            for file_info in z.infolist():
                if file_info.filename.startswith("xl/media/"):
                    filename = Path(file_info.filename).name
                    target_path = Path(output_dir) / filename
                    
                    with z.open(file_info) as source, open(target_path, 'wb') as target:
                        shutil.copyfileobj(source, target)
                    
                    extracted_files.append(str(target_path))
                    count += 1
        
        return count, extracted_files, None
    except Exception as e:
        return 0, [], str(e)

def extract_brand(file_stem):
    """Extract brand name from filename using pattern matching"""
    parts = re.split(r'[ _\-]', file_stem)
    return re.sub(r'[^a-z0-9]', '', parts[0].lower()) if parts else "unknown"

def erkenne_marken_aus_ordnern(input_folder):
    """Recognize brands from subfolders (original method)"""
    try:
        return sorted([
            f for f in os.listdir(input_folder)
            if os.path.isdir(os.path.join(input_folder, f)) and not f.startswith('.')
        ])
    except Exception as e:
        st.error(f"Error reading folders: {str(e)}")
        return []

def get_files_by_marke(input_folder, marken):
    """Get files per brand from folders"""
    spalteninhalte = {marke: [] for marke in marken}
    
    for marke in marken:
        folder_path = os.path.join(input_folder, marke)
        if os.path.isdir(folder_path):
            try:
                dateien = [
                    os.path.join(folder_path, f)
                    for f in os.listdir(folder_path)
                    if not f.startswith(".") and os.path.splitext(f)[1].lower() in ALL_ALLOWED_EXTENSIONS
                ]
                spalteninhalte[marke] = dateien
            except Exception as e:
                st.warning(f"Error reading files from {marke}: {str(e)}")
                spalteninhalte[marke] = []
    
    return spalteninhalte

def get_asset_cell(file_path, filename, col_count):
    """Create optimized asset cell for PDF with proper transparency handling"""
    ext = Path(file_path).suffix.lower()
    styles = getSampleStyleSheet()
    
    try:
        if ext in ALLOWED_IMAGE_EXTENSIONS:
            with PILImage.open(file_path) as img:
                # Preserve transparency if available
                if img.mode in ('RGBA', 'LA') or (img.mode == 'P' and 'transparency' in img.info):
                    # Create a white background for the PDF (transparency doesn't work well in PDF)
                    background = PILImage.new('RGB', img.size, (255, 255, 255))
                    if img.mode == 'RGBA':
                        # Paste the image onto the white background using the alpha channel as mask
                        background.paste(img, mask=img.split()[-1])
                    else:
                        # For other transparent modes, convert to RGBA first
                        img_rgba = img.convert('RGBA')
                        background.paste(img_rgba, mask=img_rgba.split()[-1])
                    img = background.convert('RGB')
                else:
                    # Convert non-transparent images to RGB
                    img = img.convert('RGB')
                
                orig_width, orig_height = img.size
                available_width = A4[0] - 40
                max_width = available_width / col_count
                max_height = 120

                aspect_ratio = orig_width / orig_height
                width = max_width
                height = width / aspect_ratio

                if height > max_height:
                    height = max_height
                    width = height * aspect_ratio

                buffer = io.BytesIO()
                img.save(buffer, format='PNG')
                buffer.seek(0)
                
                rl_img = RLImage(buffer, width=width, height=height)

                frame = KeepInFrame(max_width, max_height + 30, content=[
                    rl_img,
                    Paragraph(filename, styles['Normal'])
                ], hAlign='CENTER')

                return frame
        else:
            return Paragraph(f"{filename}", styles['Normal'])
    except Exception as e:
        return Paragraph(f"[Image Error]<br/>{filename}", styles['Normal'])
    

def build_marken_einzelseiten(alle_dateien_pro_marke, styles, gesamtbreite):
    """Build individual brand overview pages (original method)"""
    elements = []
    max_bildhoehe = 100

    for marke, dateien in alle_dateien_pro_marke.items():
        elements.append(Paragraph(f"Brand Overview: {marke}", styles['Heading2']))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"<b>Total Assets:</b> {len(dateien)}", styles['Normal']))
        elements.append(Spacer(1, 12))

        if not dateien:
            elements.append(Paragraph("No assets found in this brand folder.", styles['Normal']))
           # elements.append(PageBreak())
            continue

        table_data = []
        row_data = []
        spaltenbreite = gesamtbreite / 3

        for i, file_path in enumerate(dateien):
            ext = os.path.splitext(file_path)[1].lower()
            filename = os.path.basename(file_path)

            if ext in ALLOWED_IMAGE_EXTENSIONS:
                try:
                    pil_img = PILImage.open(file_path)
                    width, height = pil_img.size
                    aspect = height / width
                    new_width = spaltenbreite * 0.9
                    new_height = min(new_width * aspect, max_bildhoehe)
                    new_width = new_height / aspect

                    img = RLImage(file_path, width=new_width, height=new_height)
                    cell = [img, Spacer(1, 4), Paragraph(filename, styles['Normal'])]
                except Exception as e:
                    cell = Paragraph(f"{filename} (Error loading image)", styles['Normal'])
            else:
                cell = Paragraph(f"{filename}", styles['Normal'])

            row_data.append(cell)
            if len(row_data) == 3:
                table_data.append(row_data)
                row_data = []

        if row_data:
            while len(row_data) < 3:
                row_data.append("")
            table_data.append(row_data)

        if table_data:
            table = Table(table_data, colWidths=[spaltenbreite] * 3)
            table.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))
            elements.append(table)
        
        elements.append(PageBreak())
    
    return elements

def get_excel_worksheets_simple(excel_file_path):
    """Get list of worksheet names from Excel file - simplified version"""
    try:
        file_ext = Path(excel_file_path).suffix.lower()
        
        if file_ext == '.xlsx':
            try:
                workbook = openpyxl.load_workbook(excel_file_path, read_only=True)
                worksheets = workbook.sheetnames
                workbook.close()
                return worksheets
            except Exception:
                return ["Sheet1"]  # Fallback
                        
        elif file_ext == '.xls':
            try:
                import xlrd
                workbook = xlrd.open_workbook(excel_file_path)
                return workbook.sheet_names()
            except ImportError:
                return ["Sheet1"]  # Fallback instead of "All worksheets"
            except Exception:
                return ["Sheet1"]
        
        return []
        
    except Exception:
        return []

def add_brand_pages(elements, marken_spalten, renamed_files_by_folder_and_marke):
    """Add brand overview pages with optimized layout"""
    styles = getSampleStyleSheet()
    for nummer, marke in marken_spalten:
        assets = []
        for folder in renamed_files_by_folder_and_marke:
            assets.extend(renamed_files_by_folder_and_marke[folder].get(marke, []))

        if not assets:
            continue

        elements.append(PageBreak())
        elements.append(Paragraph(f"<b>Brand Overview: {nummer} – {marke}</b>", styles['Heading2']))
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(f"Number of assets: {len(assets)}", styles['Normal']))
        elements.append(Spacer(1, 10))

        headers = ["Asset"] * 4
        data = [headers]
        row = []
        for i, (pfad, name) in enumerate(assets):
            cell = get_asset_cell(pfad, name, 4)
            row.append(cell)
            if len(row) == 4:
                data.append(row)
                row = []
        if row:
            row.extend([""] * (4 - len(row)))
            data.append(row)

        table = Table(data, colWidths=(A4[0] - 40) / 4)
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(table)

def analyze_files_by_filename(input_folder):
    """Analyze files by extracting brand names from filenames"""
    dateien = []
    marken_set = set()
    renamed_files_by_folder_and_marke = defaultdict(lambda: defaultdict(list))

    for subfolder in sorted(Path(input_folder).iterdir()):
        if not subfolder.is_dir():
            continue

        for file in sorted(subfolder.iterdir()):
            if file.suffix.lower() not in ALL_ALLOWED_EXTENSIONS:
                continue
            marke = extract_brand(file.stem)
            marken_set.add(marke)
            dateien.append({
                "original_file": file,
                "original_folder": subfolder.name,
                "marke": marke
            })

    for eintrag in dateien:
        orig = eintrag["original_file"]
        marke = eintrag["marke"]
        folder_name = eintrag["original_folder"]
        renamed_files_by_folder_and_marke[folder_name][marke].append((orig, orig.name))

    return marken_set, renamed_files_by_folder_and_marke, dateien

def generate_pdf_report(input_folder, erste_marke=None):
    """Generate the complete PDF report (original method)"""
    styles = getSampleStyleSheet()
    gesamtbreite = 480

    marken = erkenne_marken_aus_ordnern(input_folder)
    if not marken:
        return None, "No brand folders found in the uploaded zip file."

    # Reorder brands if specified
    if erste_marke and erste_marke in marken:
        marken = [erste_marke] + [m for m in marken if m != erste_marke]

    dateien_pro_marke = get_files_by_marke(input_folder, marken)
    total_all = sum(len(dateien_pro_marke[marke]) for marke in marken)

    alle_elements = []

    # Overview page
    alle_elements.append(Paragraph("Brand Assets Overview", styles['Title']))
    alle_elements.append(Spacer(1, 20))
    
    übersicht_text = ", ".join([f"{marke}: {len(dateien_pro_marke[marke])}" for marke in marken])
    alle_elements.append(Paragraph(f"<b>Assets per brand:</b> {übersicht_text}", styles['Normal']))
    alle_elements.append(Spacer(1, 12))
    alle_elements.append(Paragraph(f"<b>Total assets:</b> {total_all}", styles['Normal']))
    alle_elements.append(PageBreak())

    # Individual brand pages
    markenseiten = build_marken_einzelseiten(dateien_pro_marke, styles, gesamtbreite)
    alle_elements.extend(markenseiten)

    # Generate PDF
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
    
    try:
        doc.build(alle_elements)
        pdf_buffer.seek(0)
        return pdf_buffer, None
    except Exception as e:
        return None, f"Error generating PDF: {str(e)}"
def extract_zip_to_temp(uploaded_file):
    """Extract uploaded zip file to temporary directory"""
    temp_dir = tempfile.mkdtemp()
    with zipfile.ZipFile(uploaded_file, 'r') as zip_ref:
        zip_ref.extractall(temp_dir)
    return temp_dir

def extract_brand(file_stem):
    """Extract brand name from filename using pattern matching"""
    parts = re.split(r'[ _\-]', file_stem)
    return re.sub(r'[^a-z0-9]', '', parts[0].lower()) if parts else "unknown"

def clean_letters_only(text):
    return re.sub(r'[^A-Za-z]', '', text)

def get_cleaned_filename_without_brand(filename, brand):
    full_letters = clean_letters_only(Path(filename).stem.lower())
    return full_letters.replace(brand, '', 1)

def generate_filename_based_pdf_report(input_folder, erste_marke=None, processed_files_mapping=None):
    """Generate PDF report based on filename brand analysis with final processed labels"""
    marken_set, renamed_files_by_folder_and_marke, all_files = analyze_files_by_filename(input_folder)
    
    if not marken_set:
        return None, "No brands found in filenames."

    # Create brand index
    marken_index = {}
    if erste_marke and erste_marke in marken_set:
        marken_index[erste_marke] = "01"
        aktuelle_nummer = 2
        for marke in sorted(marken_set):
            if marke != erste_marke:
                marken_index[marke] = f"{aktuelle_nummer:02d}"
                aktuelle_nummer += 1
    else:
        for i, marke in enumerate(sorted(marken_set), 1):
            marken_index[marke] = f"{i:02d}"

    # Generate PDF
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, leftMargin=20, rightMargin=20, topMargin=40, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    nummer_zu_marke = {v: k for k, v in marken_index.items()}
    marken_spalten = sorted(nummer_zu_marke.items())

    # Process each folder with final processed names
    for folder in sorted(renamed_files_by_folder_and_marke):
        elements.append(Paragraph(f"<b>Folder: {folder}</b>", styles['Heading2']))

        abschnitt = renamed_files_by_folder_and_marke[folder]
        total = 0
        lines = []
        for nummer, marke in marken_spalten:
            count = len(abschnitt.get(marke, []))
            total += count
            lines.append(f"{nummer} ({marke}): {count}")
        lines.append(f"Total: {total}")
        elements.append(Paragraph("<br/>".join(lines), styles['Normal']))
        elements.append(Spacer(1, 10))

        headers = [f"{nummer} ({marke})" for nummer, marke in marken_spalten]
        data = [headers]
        col_data = []
        max_rows = 0
        
        for _, marke in marken_spalten:
            eintraege = abschnitt.get(marke, [])
            zellen = []
            for original_path, original_name in eintraege:
                # Generate the final processed ID name
                blocknummer = re.sub(r'\D', '', folder)[:2].zfill(2)
                markennummer = marken_index[marke]
                cleaned = get_cleaned_filename_without_brand(original_name, marke)
                # Create the final ID without file extension for display
                final_id = f"{markennummer}B{blocknummer}{marke}{cleaned}"
                
                # Create cell with original path but final ID label
                cell = get_asset_cell(original_path, final_id, len(headers))
                zellen.append(cell)
            col_data.append(zellen)
            max_rows = max(max_rows, len(zellen))

        for i in range(max_rows):
            row = []
            for col in col_data:
                row.append(col[i] if i < len(col) else "")
            data.append(row)

        col_width = (A4[0] - 40) / len(headers)
        t = Table(data, colWidths=[col_width] * len(headers))
        t.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(t)
        elements.append(PageBreak())

    # Total overview
    elements.append(Paragraph("<b>Total Overview</b>", styles['Heading2']))
    global_counts = defaultdict(int)
    for folder_data in renamed_files_by_folder_and_marke.values():
        for marke, daten in folder_data.items():
            global_counts[marke] += len(daten)
    
    gesamt = 0
    summary = []
    for nummer, marke in marken_spalten:
        count = global_counts[marke]
        summary.append(f"{marke}: {count} assets")
        gesamt += count
    summary.append(f"Total number of all assets: {gesamt}")
    elements.append(Paragraph("<br/>".join(summary), styles['Normal']))

    # Add brand pages with final processed labels
    add_brand_pages_with_final_labels(elements, marken_spalten, renamed_files_by_folder_and_marke, marken_index)

    try:
        doc.build(elements)
        pdf_buffer.seek(0)
        return pdf_buffer, None
    except Exception as e:
        return None, f"Error generating PDF: {str(e)}"

def add_brand_pages_with_final_labels(elements, marken_spalten, renamed_files_by_folder_and_marke, marken_index):
    """Add brand overview pages with final processed ID labels"""
    styles = getSampleStyleSheet()
    for nummer, marke in marken_spalten:
        assets = []
        for folder in renamed_files_by_folder_and_marke:
            folder_assets = renamed_files_by_folder_and_marke[folder].get(marke, [])
            for original_path, original_name in folder_assets:
                # Generate the final processed ID name (same logic as in Excel generation)
                blocknummer = re.sub(r'\D', '', folder)[:2].zfill(2)
                markennummer = marken_index[marke]
                cleaned = get_cleaned_filename_without_brand(original_name, marke)
                final_id = f"{markennummer}B{blocknummer}{marke}{cleaned}"
                assets.append((original_path, final_id))

        if not assets:
            continue

        elements.append(PageBreak())
        elements.append(Paragraph(f"<b>Brand Overview: {nummer} – {marke}</b>", styles['Heading2']))
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(f"Number of assets: {len(assets)}", styles['Normal']))
        elements.append(Spacer(1, 10))

        headers = ["Asset"] * 4
        data = [headers]
        row = []
        for i, (original_path, final_id) in enumerate(assets):
            # Use final processed ID as the label
            cell = get_asset_cell(original_path, final_id, 4)
            row.append(cell)
            if len(row) == 4:
                data.append(row)
                row = []
        if row:
            row.extend([""] * (4 - len(row)))
            data.append(row)

        table = Table(data, colWidths=(A4[0] - 40) / 4)
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(table)

def generate_two_section_pdf_report(input_folder, erste_marke=None):
    """Generate PDF report with two sections: by block and by brand - WITH EXTENSIONS"""
    marken_set, renamed_files_by_folder_and_marke, all_files = analyze_files_by_filename(input_folder)
    
    if not marken_set:
        return None, "No brands found in filenames."

    # Create brand index
    marken_index = {}
    if erste_marke and erste_marke in marken_set:
        marken_index[erste_marke] = "01"
        aktuelle_nummer = 2
        for marke in sorted(marken_set):
            if marke != erste_marke:
                marken_index[marke] = f"{aktuelle_nummer:02d}"
                aktuelle_nummer += 1
    else:
        for i, marke in enumerate(sorted(marken_set), 1):
            marken_index[marke] = f"{i:02d}"

    # Generate PDF
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, leftMargin=20, rightMargin=20, topMargin=40, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    nummer_zu_marke = {v: k for k, v in marken_index.items()}
    marken_spalten = sorted(nummer_zu_marke.items())

    # SECTION 1: BY BLOCK/FOLDER
    elements.append(Paragraph("<b>Section 1: Assets by Block</b>", styles['Title']))
    elements.append(Spacer(1, 20))

    for folder in sorted(renamed_files_by_folder_and_marke):
        elements.append(Paragraph(f"<b>Folder: {folder}</b>", styles['Heading2']))

        abschnitt = renamed_files_by_folder_and_marke[folder]
        total = 0
        lines = []
        for nummer, marke in marken_spalten:
            count = len(abschnitt.get(marke, []))
            total += count
            lines.append(f"{nummer} ({marke}): {count}")
        lines.append(f"Total: {total}")
        elements.append(Paragraph("<br/>".join(lines), styles['Normal']))
        elements.append(Spacer(1, 10))

        headers = [f"{nummer} ({marke})" for nummer, marke in marken_spalten]
        data = [headers]
        col_data = []
        max_rows = 0
        
        for _, marke in marken_spalten:
            eintraege = abschnitt.get(marke, [])
            zellen = []
            for p, original_name in eintraege:
                # Generate the ID name - KEEP THE FILE EXTENSION
                blocknummer = re.sub(r'\D', '', folder)[:2].zfill(2)
                markennummer = marken_index[marke]
                cleaned = get_cleaned_filename_without_brand(original_name, marke)
                # Keep the original file extension
                file_extension = Path(original_name).suffix.lower()
                id_name = f"{markennummer}B{blocknummer}{marke}{cleaned}{file_extension}"
                
                # Pass the full filename with extension to get_asset_cell
                cell = get_asset_cell(p, id_name, len(headers))
                zellen.append(cell)
            col_data.append(zellen)
            max_rows = max(max_rows, len(zellen))

        for i in range(max_rows):
            row = []
            for col in col_data:
                row.append(col[i] if i < len(col) else "")
            data.append(row)

        col_width = (A4[0] - 40) / len(headers)
        t = Table(data, colWidths=[col_width] * len(headers))
        t.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(t)
        elements.append(PageBreak())

    # SECTION 2: BY BRAND
    elements.append(Paragraph("<b>Section 2: Assets by Brand</b>", styles['Title']))
    elements.append(Spacer(1, 20))

    # Total overview first
    elements.append(Paragraph("<b>Total Overview</b>", styles['Heading2']))
    global_counts = defaultdict(int)
    for folder_data in renamed_files_by_folder_and_marke.values():
        for marke, daten in folder_data.items():
            global_counts[marke] += len(daten)
    
    gesamt = 0
    summary = []
    for nummer, marke in marken_spalten:
        count = global_counts[marke]
        summary.append(f"{marke}: {count} assets")
        gesamt += count
    summary.append(f"Total number of all assets: {gesamt}")
    elements.append(Paragraph("<br/>".join(summary), styles['Normal']))
    elements.append(PageBreak())
    first_brand = True
    # Individual brand pages
    for nummer, marke in marken_spalten:
        assets = []
        for folder in renamed_files_by_folder_and_marke:
            folder_assets = renamed_files_by_folder_and_marke[folder].get(marke, [])
            brand_counter = 1  # Start counting from 1 for each brand in each folder
            for pfad, original_name in folder_assets:
                # Generate the ID name with sequential count - KEEP THE FILE EXTENSION
                blocknummer = re.sub(r'\D', '', folder)[:2].zfill(2)
                markennummer = marken_index[marke]
                count_str = f"{brand_counter:02d}"
                cleaned = get_cleaned_filename_without_brand(original_name, marke)
                # Keep the original file extension
                file_extension = Path(original_name).suffix.lower()
                id_name = f"{markennummer}B{blocknummer}{marke}{count_str}{cleaned}{file_extension}"
                assets.append((pfad, id_name))
                brand_counter += 1

        if not assets:
            continue



        # Inside the loop for each brand
        if not first_brand:
            elements.append(PageBreak())
        else:
            first_brand = False
        elements.append(Paragraph(f"<b>Brand Overview: {nummer} – {marke}</b>", styles['Heading2']))
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(f"Assets per Brand: {', '.join([f'{k}: {len([a for a in assets if extract_brand(Path(a[1]).stem) == k])}' for k in [marke]])}", styles['Normal']))
        elements.append(Paragraph(f"Total assets: {len(assets)}", styles['Normal']))
        elements.append(Spacer(1, 10))
        

        headers = ["Asset"] * 4
        data = [headers]
        row = []
        for i, (pfad, id_name) in enumerate(assets):
            # Pass the full filename with extension to get_asset_cell
            cell = get_asset_cell(pfad, id_name, 4)
            row.append(cell)
            if len(row) == 4:
                data.append(row)
                row = []
        if row:
            row.extend([""] * (4 - len(row)))
            data.append(row)

        table = Table(data, colWidths=(A4[0] - 40) / 4)
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(table)

    try:
        doc.build(elements)
        pdf_buffer.seek(0)
        return pdf_buffer, None
    except Exception as e:
        return None, f"Error generating PDF: {str(e)}"
